# -*- coding: utf-8 -*-
"""오염 안 된 홀드아웃 원재료 수집 — 정답셋에 한 번도 안 쓰인 상품에서만 뽑는다.

**왜 상품 단위로 가르는가**: "안 읽은 이미지"와 "안 쓴 상품"은 다르다. 같은 상세페이지는
슬로건·성분표·인증 문구가 이미지마다 반복되므로, 이미 튜닝에 쓴 상품의 다른 이미지에서
뽑은 문장은 "처음 보는 문장"이라는 보장이 없다. 규칙집을 그 상품 문장으로 고쳤기 때문이다.
(2026-08-20 실측: 미OCR 222장 중 190장이 이미 라벨 있는 20개 상품의 것이었다.)

    ./venv/bin/python scripts/build_clean_holdout.py --stage1
"""
import argparse
import json
import re
import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
for _p in (ROOT, ROOT / "src"):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

import openpyxl  # noqa: E402

from barum.pipeline import _ocr_image  # noqa: E402  운영과 같은 경로(타일 분할 후 OCR)
from barum.vlm import get_vlm  # noqa: E402

DETAILS = ROOT / "11st_probe_cosmetic" / "details"
ANSWER_KEY = ROOT / "11st_probe_cosmetic" / "read_test" / "_combined_answer_key.json"
LABELSET = ROOT / "11st_probe_cosmetic" / "read_test" / "label_worksheet_combined.xlsx"
OUT = ROOT / "data" / "clean_holdout_raw.jsonl"

# 2026-08-20 전성분 추출로 읽은 이미지(평가셋 7개 상품). 여기도 '이미 읽음'이다.
_INGREDIENT_PASS = ["1010944945", "1403306051", "24500688", "24505724",
                    "7628382624", "8783520869", "9126459148"]


def _norm(s: str) -> str:
    """중복 판정용 정규화 — 공백·문장부호를 지운 형태로 비교한다."""
    return re.sub(r"[\s\W_]+", "", s)


def _already_read() -> set[tuple[str, str]]:
    done = set()
    for e in json.loads(ANSWER_KEY.read_text(encoding="utf-8")):
        parts = e.get("png", "").split("_", 2)
        if len(parts) == 3:
            done.add((parts[1], Path(parts[2]).stem))
    for code in _INGREDIENT_PASS:
        imgs = sorted(p for p in (DETAILS / code).iterdir()
                      if p.suffix.lower() in (".jpg", ".png", ".gif"))
        for p in imgs[-3:]:
            done.add((code, p.stem))
        if code == "1010944945":
            for p in imgs[:-3]:
                done.add((code, p.stem))
    return done


def _labelset() -> tuple[set[str], set[str]]:
    """(라벨이 있는 상품코드, 정규화된 기존 문장)."""
    wb = openpyxl.load_workbook(LABELSET)
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    i = {h: n for n, h in enumerate(hdr)}
    products, sents = set(), set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[i["문장"]]:
            continue
        products.add(str(r[i["상품코드"]]))
        sents.add(_norm(r[i["문장"]]))
    return products, sents


def clean_products() -> dict[str, list[Path]]:
    """정답셋에 한 번도 안 나온 상품의, 아직 안 읽은 이미지."""
    done = _already_read()
    labeled, _ = _labelset()
    out = {}
    for d in sorted(DETAILS.iterdir()):
        if not d.is_dir() or d.name in labeled:
            continue
        imgs = [p for p in sorted(d.iterdir())
                if p.suffix.lower() in (".jpg", ".png", ".gif") and (d.name, p.stem) not in done]
        if imgs:
            out[d.name] = imgs
    return out


def main() -> None:
    pool = clean_products()
    n_img = sum(len(v) for v in pool.values())
    print(f"오염 안 된 상품 {len(pool)}개 / 이미지 {n_img}장\n", flush=True)

    _, known = _labelset()
    vlm = get_vlm("gemini")
    rows, seen = [], set()
    dup_known = dup_self = 0
    t0 = time.perf_counter()

    for code, imgs in pool.items():
        got = 0
        for p in imgs:
            try:
                # **운영 경로로 읽는다.** 통짜 이미지를 한 번에 보내면 긴 페이지에서
                # 응답 JSON이 잘려 통째로 실패한다(2026-08-20 실측: 4400px 넘는 3장이
                # 반복 실패, 타일 경로로는 3장 다 성공). 운영 `/check`도 타일로 자른다.
                sents, _n_failed = _ocr_image(p.read_bytes(), p.name, vlm, verbose=False)
            except Exception as e:
                # 과금 호출은 재시도하지 않는다(CLAUDE.md §E). 실패로 기록하고 넘어간다.
                print(f"  [실패] {code}/{p.name}: {type(e).__name__}")
                continue
            for s in sents:
                t = (s.get("text") or "").strip()
                if not t:
                    continue
                n = _norm(t)
                if not n:
                    continue
                if n in known:      # 963셋에 이미 있는 문장 = 오염
                    dup_known += 1
                    continue
                if n in seen:       # 같은 페이지 안 반복
                    dup_self += 1
                    continue
                seen.add(n)
                rows.append({"product": code, "image": p.name, "text": t})
                got += 1
        print(f"  {code}: 이미지 {len(imgs)}장 → 새 문장 {got}개", flush=True)

    OUT.write_text("\n".join(json.dumps(r, ensure_ascii=False) for r in rows), encoding="utf-8")
    print(f"\n=== 1단계 결과 ({time.perf_counter() - t0:.0f}초, 토큰 {getattr(vlm, 'total_tokens', 0):,}) ===")
    print(f"새 문장          : {len(rows)}개")
    print(f"963셋과 중복 제거 : {dup_known}개")
    print(f"페이지 내 중복 제거: {dup_self}개")
    print(f"저장: {OUT}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--stage1", action="store_true", help="깨끗한 상품만 OCR (기본 동작)")
    ap.parse_args()
    main()
