"""정답셋(label_worksheet.xlsx) 검수 — 프로덕션과 다른 고성능 LLM 초안 + 사람 최종확인.

대수(팀원B)가 1차로 매긴 240문장 잠정 라벨을, 판정 파이프라인이 쓰는 gpt-5-mini와는
다른 모델(gpt-5.6-luna)로 다시 매겨 초안+근거를 낸다. 같은 모델로 검수하면 그 모델이
스스로 낸 정답으로 스스로를 채점하는 순환참조가 생겨 탐지율이 부풀려지므로, 반드시
프로덕션과 다른 모델을 쓴다(하니 지시, PROJECT.md 참고).

이 스크립트는 정답을 확정하지 않는다. LLM 초안 라벨·근거를 기존 대수 라벨 옆에
나란히 써서 사람이 불일치 행 위주로 최종 확인하기 쉽게 만드는 게 목적이다.

실행(backend/에서):
    python scripts/review_labelset.py                 # 전체 240문장 실행 (API 호출)
    python scripts/review_labelset.py --dry            # 호출 없이 배선만 점검
    python scripts/review_labelset.py --limit 2        # 이미지 2장만(스모크 테스트)

출력:
    11st_probe_cosmetic/read_test/label_worksheet_reviewed.xlsx
"""
import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, "src")
sys.path.insert(0, ".")  # tile_split.py (backend 루트)
from barum.models import JudgmentFlag, ViolationType  # noqa: E402
from barum.reference.context import build_judgment_context  # noqa: E402
from barum.vlm import get_vlm  # noqa: E402

_READ_TEST = Path("11st_probe_cosmetic/read_test")
_IMAGES_DIR = _READ_TEST / "images"
_ANSWER_KEY = _READ_TEST / "_llm_answer_key.json"
_LABEL_XLSX = _READ_TEST / "label_worksheet.xlsx"
_OUT_XLSX = _READ_TEST / "label_worksheet_reviewed.xlsx"
_MISMATCH_XLSX = _READ_TEST / "label_worksheet_mismatches.xlsx"

# 불일치 중 recall 우선 원칙상 가장 위험한 방향(사람=위반→LLM=합법 등)이라 비비가
# 최우선으로 봐야 하는 (이미지번호, 문장#). 2026-08-13 검수 실행 결과 기준.
_PRIORITY_CASES = {("13", 20)}

# 결론이 "그 성분이 실제로 기준 함량을 채웠는가"에 달린 불일치(2호_기능성오인 계열).
# 문장·이미지만으로는 함량을 확정할 수 없으니, 하니 지시대로 비비가 따로 안 찾아보고
# 바로 검토필요로 두게 표시만 해준다(PM 2026-08-13 추가 규칙). 판정 자체을 여기서
# 대신 내리는 게 아니라 "이건 함량 확인이 필요한 유형이다"라는 분류 힌트일 뿐이라,
# 최종 기입은 여전히 비비 몫이다. 36건 원문·근거를 읽고 직접 골랐다(2026-08-13).
_AMOUNT_DEPENDENT_CASES = {
    ("06", 6),   # 안나홀츠 안티링클 아이크림 — 주름개선 표방, 고시원료·함량 확인 필요
    ("07", 1),   # 홀츠포맨 콜라겐 안티링크 화이트닝 올인원 — 미백·주름개선 표방
    ("07", 3),   # 미백, 주름개선 2중 기능성 — LLM 근거가 직접 "기준함량 정합성" 언급
    ("07", 4),   # 콜라겐 추출물 1000ppm 함유 — LLM 근거가 직접 "함량 산정 기준" 언급
    ("09", 83),  # 피부의 미백에 도움을 준다 — 미백 표방, 고시원료·함량 확인 필요
    ("09", 84),  # 피부의 주름개선에 도움을 준다 — 주름개선 표방, 고시원료·함량 확인 필요
    ("13", 20),  # 미백ㆍ주름개선 이중기능성 화장품 — 최우선(_PRIORITY_CASES)과 겹침
}

# 검수용 모델 — 프로덕션(gpt-5-mini)과 다른 모델이어야 순환참조가 안 생긴다.
REVIEW_MODEL = "gpt-5.6-luna"

# label_worksheet.xlsx "판정" 컬럼이 실제로 쓰는 라벨 6종(models.py ViolationType +
# JudgmentFlag를 사람이 보기 편하게 합친 것). "애매"는 사람 전용(불확실 보류) 값이라
# LLM에는 요구하지 않는다 — LLM은 항상 이 중 하나로 커밋한다(불확실하면 검토필요).
_VIOLATION_TYPES = [
    ViolationType.type_1_drug_misperception.value,
    ViolationType.type_2_functional_misperception.value,
    ViolationType.type_5_deception.value,
]
_LABELS = [ViolationType.legal.value, *_VIOLATION_TYPES, ViolationType.out_of_scope.value, JudgmentFlag.needs_review.value]

REVIEW_PROMPT_TEMPLATE = """너는 한국 화장품 광고 위반 판정 전문가다. 아래 [판정 근거]를 참고해서,
이 이미지에서 사람이 이미 확정한 아래 문장들 각각을 판정하라. 문장 자체는 이미 확정됐으니
새로 읽지 말고(이미지에 없는 문장이 섞여 있어도 그대로 판정), 목록 순서 그대로 라벨과 근거만 매겨라.

[판정 근거]
{context}

[문장 목록]
{sentence_list}

라벨(정확히 이 중 하나): 합법 / 1호_의약품오인 / 2호_기능성오인 / 5호_거짓과장기만 / 대상외 / 검토필요
- 미탐(위반을 합법으로 놓침)이 제일 나쁜 실수다. 확신 없으면 검토필요로, 안전 쪽으로 판단하라.
- 대상외는 성분명 나열·브랜드명·인증마크·거래정보 등 광고 문구가 아닌 것.

JSON으로만 답하라:
{{"results": [{{"문장번호": 1, "label": "라벨", "reason": "짧은 근거"}}, ...]}}
문장번호는 반드시 [문장 목록]의 번호와 1:1로 대응해야 하고, 개수도 정확히 같아야 한다."""


def load_worksheet_rows() -> list[dict]:
    """label_worksheet.xlsx "라벨링" 시트를 읽는다.

    반환: [{row_idx(엑셀 행번호), nn, code, seq, sentence, label, vtype}, ...]
    row_idx를 들고 있어야 출력 때 원본 행에 그대로 이어붙일 수 있다.
    """
    wb = openpyxl.load_workbook(_LABEL_XLSX)
    ws = wb["라벨링"]
    rows = []
    for r in range(2, ws.max_row + 1):
        nn = str(ws.cell(r, 1).value or "").strip()
        if not nn:
            continue
        rows.append({
            "row_idx": r,
            "nn": nn,
            "code": str(ws.cell(r, 2).value or "").strip(),
            "seq": ws.cell(r, 3).value,
            "sentence": str(ws.cell(r, 4).value or "").strip(),
            "human_label": str(ws.cell(r, 5).value or "").strip(),
            "human_vtype": str(ws.cell(r, 6).value or "").strip(),
        })
    return rows


def load_image_map() -> dict[str, str]:
    """_llm_answer_key.json에서 이미지번호(nn) -> png 파일명을 읽는다."""
    meta = json.loads(_ANSWER_KEY.read_text(encoding="utf-8"))
    return {m["nn"]: m["png"] for m in meta}


def _build_prompt(context: str, sentences: list[str]) -> str:
    numbered = "\n".join(f"{i}. {s}" for i, s in enumerate(sentences, 1))
    return REVIEW_PROMPT_TEMPLATE.format(context=context, sentence_list=numbered)


def review_image(nn: str, rows: list[dict], image_bytes: bytes, vlm, context: str) -> dict[int, dict]:
    """한 이미지(nn)에 속한 문장들을 한 번에 판정받는다.

    반환: {문장번호(1-based, rows 순서): {label, reason}}. 호출 실패·개수 불일치는
    빈 dict(호출자가 "미판정"으로 기록, 재시도하지 않는다 — 과금 호출).
    """
    sentences = [r["sentence"] for r in rows]
    prompt = _build_prompt(context, sentences)
    try:
        res = vlm.generate_json(prompt, [image_bytes])
    except Exception as e:
        print(f"    [skip] {nn}: 호출 실패 {type(e).__name__}: {e}")
        return {}

    raw = res.get("results", []) if isinstance(res, dict) else []
    if len(raw) != len(sentences):
        print(f"    [skip] {nn}: 응답 {len(raw)}개 != 문장 {len(sentences)}개, 미판정으로 남김")
        return {}

    out = {}
    for item in raw:
        try:
            n = int(item.get("문장번호"))
        except (TypeError, ValueError):
            continue
        label = (item.get("label") or "").strip()
        if label not in _LABELS:
            print(f"    [경고] {nn} #{n}: 규격 밖 라벨 '{label}', 미판정으로 남김")
            continue
        out[n] = {"label": label, "reason": (item.get("reason") or "").strip()}
    return out


def _label_to_worksheet_columns(label: str) -> tuple[str, str]:
    """review_image의 라벨(6종)을 worksheet 스키마(판정/위반유형 두 컬럼)로 나눈다."""
    if label in _VIOLATION_TYPES:
        return JudgmentFlag.violation.value, label
    return label, ""  # 합법 / 대상외 / 검토필요는 그대로, 위반유형 칸은 비움


def write_reviewed_xlsx(rows: list[dict], results: dict[str, dict[int, dict]]) -> None:
    """원본 워크북을 복사해 LLM 초안 컬럼(H~K)을 덧붙인다. 원본 대수 라벨(E/F)은 안 건드림."""
    wb = openpyxl.load_workbook(_LABEL_XLSX)
    ws = wb["라벨링"]

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    mismatch_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ambiguous_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin = Border(*(Side(style="thin"),) * 4)
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = ["LLM_판정", "LLM_위반유형", "LLM_근거", "일치여부", "비비_최종판단"]
    for ci, h in enumerate(headers, 8):  # H=8
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.border = header_font, header_fill, thin
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 50
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 40  # 비비가 인용 검증 후 채우는 감사추적용 칸, 빈 채로 둔다

    by_nn = defaultdict(list)
    for r in rows:
        by_nn[r["nn"]].append(r)

    n_match = n_mismatch = n_unjudged = n_human_ambiguous = 0

    for nn, nn_rows in by_nn.items():
        drafts = results.get(nn, {})
        for i, r in enumerate(nn_rows, 1):
            row_idx = r["row_idx"]
            draft = drafts.get(i)
            if draft is None:
                ws.cell(row=row_idx, column=8, value="(미판정)").border = thin
                ws.cell(row=row_idx, column=9, value="").border = thin
                ws.cell(row=row_idx, column=10, value="").border = thin
                ws.cell(row=row_idx, column=11, value="").border = thin
                n_unjudged += 1
                continue

            llm_label, llm_vtype = _label_to_worksheet_columns(draft["label"])
            human_label = r["human_label"]

            if human_label == "애매":
                match_text = "사람애매"
                fill = ambiguous_fill
                n_human_ambiguous += 1
            elif human_label == llm_label:
                match_text = "일치"
                fill = None
                n_match += 1
            else:
                match_text = "불일치"
                fill = mismatch_fill
                n_mismatch += 1

            c8 = ws.cell(row=row_idx, column=8, value=llm_label)
            c9 = ws.cell(row=row_idx, column=9, value=llm_vtype)
            c10 = ws.cell(row=row_idx, column=10, value=draft["reason"])
            c11 = ws.cell(row=row_idx, column=11, value=match_text)
            for c in (c8, c9, c10, c11):
                c.border = thin
                if fill:
                    c.fill = fill
            c10.alignment = wrap

    ws.auto_filter.ref = f"A1:K{ws.max_row}"
    ws.freeze_panes = "A2"
    wb.save(_OUT_XLSX)

    total = n_match + n_mismatch + n_unjudged + n_human_ambiguous
    print(f"\n저장: {_OUT_XLSX}")
    print(
        f"일치 {n_match} / 불일치 {n_mismatch} / 사람애매 {n_human_ambiguous} / "
        f"미판정 {n_unjudged} (총 {total})"
    )
    if n_mismatch:
        print(f"불일치 {n_mismatch}건은 노란색으로 표시됨 — 사람 최종확인 우선순위.")


def ensure_final_column() -> None:
    """이미 생성된 label_worksheet_reviewed.xlsx에 비비_최종판단 컬럼(L)이 없으면 추가한다.

    API 호출 없이 기존 산출물만 패치한다 — review_labelset.py를 이 컬럼 포함 버전으로
    다시 돌려 API 비용을 또 쓸 필요 없게, 과거 산출물 호환용으로 둔다.
    """
    wb = openpyxl.load_workbook(_OUT_XLSX)
    ws = wb["라벨링"]
    if ws.cell(row=1, column=12).value == "비비_최종판단":
        print("이미 있음, 건너뜀.")
        return

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin = Border(*(Side(style="thin"),) * 4)

    c = ws.cell(row=1, column=12, value="비비_최종판단")
    c.font, c.fill, c.border = header_font, header_fill, thin
    ws.column_dimensions["L"].width = 40
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=12).border = thin
    ws.auto_filter.ref = f"A1:L{ws.max_row}"
    wb.save(_OUT_XLSX)
    print(f"비비_최종판단 컬럼(L) 추가 완료: {_OUT_XLSX}")


def export_mismatches() -> None:
    """불일치 행만 뽑아 비비에게 넘길 별도 워크북을 만든다.

    label_worksheet_reviewed.xlsx에서 "일치여부"=="불일치"인 행을, 대수 판정/근거메모와
    LLM 판정/근거를 나란히 놓아 인용 검증하기 쉽게 정리한다. _PRIORITY_CASES에 지정된
    (이미지, 문장#)는 recall 우선 원칙상 가장 위험한 방향(사람=위반→LLM=합법 등)이라
    빨간색으로 최우선 표시한다. _AMOUNT_DEPENDENT_CASES는 결론이 성분 함량 확인에
    달린 유형이라(PM 2026-08-13 규칙: 이런 건 무조건 검토필요) 초록색으로 표시해
    비비가 바로 규칙 적용하고 넘어갈 수 있게 한다 — 최종 기입은 비비 몫이다.
    """
    wb = openpyxl.load_workbook(_OUT_XLSX)
    ws = wb["라벨링"]

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "불일치"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    priority_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    amount_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    thin = Border(*(Side(style="thin"),) * 4)
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = [
        "이미지", "상품코드", "문장#", "문장",
        "대수_판정", "대수_위반유형", "대수_근거메모",
        "LLM_판정", "LLM_위반유형", "LLM_근거",
        "우선순위", "성분함량_의존", "비비_최종판단",
    ]
    widths = [8, 12, 8, 45, 10, 16, 30, 10, 16, 45, 10, 14, 40]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = out_ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.border = header_font, header_fill, thin
        out_ws.column_dimensions[c.column_letter].width = w

    row_out = 2
    n_priority = 0
    n_amount_dependent = 0
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 11).value or "").strip() != "불일치":
            continue
        nn = str(ws.cell(r, 1).value or "").strip()
        seq = ws.cell(r, 3).value
        is_priority = (nn, seq) in _PRIORITY_CASES
        is_amount_dependent = (nn, seq) in _AMOUNT_DEPENDENT_CASES
        values = [
            nn, ws.cell(r, 2).value, seq, ws.cell(r, 4).value,
            ws.cell(r, 5).value, ws.cell(r, 6).value, ws.cell(r, 7).value,
            ws.cell(r, 8).value, ws.cell(r, 9).value, ws.cell(r, 10).value,
            "최우선" if is_priority else "",
            "예 (규칙: 검토필요)" if is_amount_dependent else "",
            "",
        ]
        for ci, v in enumerate(values, 1):
            c = out_ws.cell(row=row_out, column=ci, value=v)
            c.border = thin
            if ci in (4, 7, 10):
                c.alignment = wrap
            if is_priority:
                c.fill = priority_fill
            elif is_amount_dependent:
                c.fill = amount_fill
        n_priority += is_priority
        n_amount_dependent += is_amount_dependent
        row_out += 1

    out_ws.auto_filter.ref = f"A1:M{row_out - 1}"
    out_ws.freeze_panes = "A2"
    out_wb.save(_MISMATCH_XLSX)
    print(
        f"불일치 {row_out - 2}건 저장: {_MISMATCH_XLSX} "
        f"(최우선 {n_priority}건 빨간색, 성분함량_의존 {n_amount_dependent}건 초록색 표시)"
    )


def main() -> None:
    ap = argparse.ArgumentParser(description="정답셋 검수: gpt-5.6-luna 초안 + 근거")
    ap.add_argument("--dry", action="store_true", help="API 호출 없이 배선만 점검")
    ap.add_argument("--limit", type=int, default=None, help="이미지 N장만 처리(스모크 테스트)")
    ap.add_argument(
        "--add-final-column", action="store_true",
        help="API 호출 없이 기존 산출물에 비비_최종판단 컬럼만 추가(과거 산출물 호환용)",
    )
    ap.add_argument(
        "--export-mismatches", action="store_true",
        help="API 호출 없이 기존 산출물에서 불일치 행만 별도 워크북으로 뽑는다",
    )
    args = ap.parse_args()

    if args.add_final_column:
        ensure_final_column()
        return
    if args.export_mismatches:
        export_mismatches()
        return

    print("정답셋 로드...")
    rows = load_worksheet_rows()
    image_map = load_image_map()
    by_nn = defaultdict(list)
    for r in rows:
        by_nn[r["nn"]].append(r)
    nns = sorted(by_nn.keys())
    if args.limit:
        nns = nns[: args.limit]
    print(f"  문장 {len(rows)}개, 이미지 {len(by_nn)}장 (이번 실행 대상 {len(nns)}장)")
    print(f"  검수 모델: {REVIEW_MODEL} (프로덕션 gpt-5-mini와 다른 모델, 순환참조 회피)")

    if args.dry:
        print("\n[dry run] 배선 점검 완료. API 호출 없이 종료.")
        for nn in nns:
            print(f"  [{nn}] {image_map.get(nn, '?')} / 문장 {len(by_nn[nn])}개")
        return

    vlm = get_vlm("openai", model=REVIEW_MODEL)
    context = build_judgment_context()
    print(f"  RAG 근거 블록: {len(context)}자\n")

    results: dict[str, dict[int, dict]] = {}
    for nn in nns:
        png = image_map.get(nn)
        if not png:
            print(f"  [skip] {nn}: 이미지 매핑 없음")
            continue
        img_path = _IMAGES_DIR / png
        print(f"  [{nn}] {png} ({len(by_nn[nn])}문장)...", end=" ", flush=True)
        drafts = review_image(nn, by_nn[nn], img_path.read_bytes(), vlm, context)
        results[nn] = drafts
        print(f"{len(drafts)}/{len(by_nn[nn])} 판정 완료")

    print(f"\n총 토큰: {getattr(vlm, 'total_tokens', 0)}")

    # limit로 일부만 돌렸으면 나머지는 미판정으로 남기고(rows 전체 기준으로 쓰되
    # results에 없는 nn은 write_reviewed_xlsx에서 자동으로 전부 미판정 처리됨).
    write_reviewed_xlsx(rows, results)


if __name__ == "__main__":
    main()
