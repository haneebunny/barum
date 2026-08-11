"""평가 홀드아웃 스프레드시트 2개를 만든다 (라벨러 2명 분담).

실행: venv/bin/python scripts/build_holdout.py
출력:
  data/holdout_A.xlsx, data/holdout_B.xlsx  ← 팀원 배포용 (label 칸 비어 있음)
  data/holdout_master.jsonl                 ← 내부용 (사전분류 힌트 포함, 배포 금지)

근거: 요구사항정의서 FR-23~FR-26, 라벨링_기준서 §1·§5
구글 시트 업로드를 전제로 단순한 기능(목록 검증·COUNTIF)만 쓴다.
"""

import argparse
import glob
import json
import random
import re
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

from barum.judge.prescreen import LABELS  # noqa: E402

PRESCREEN_GLOB = "prescreen*.jsonl"

# 기준서 §4.4 — 한 문장에 여러 유형이면 위해등급이 높은 것을 택한다.
# 모델 두 개의 판정을 합칠 때도 같은 순서를 쓴다(미탐 회수 = recall 우선).
SEVERITY = [
    "2호_의약품오인", "1호_질병표방", "3호_건기식오인",
    "4호_거짓과장", "5호_소비자기만", "대상외", "합법",
]

# 라벨링_기준서 §5.2 유형별 최소 확보량
QUOTA = {
    "합법": 60,
    "1호_질병표방": 40,
    "3호_건기식오인": 40,
    "4호_거짓과장": 30,
    "5호_소비자기만": 30,
    "2호_의약품오인": 20,
}
COLUMNS = [
    "id", "product_id", "product_url", "product_name", "product_type",
    "certified_function", "disclaimer", "tile_text",
    "context_before", "sentence", "context_after",
    "label", "confidence", "note", "labeler", "is_cross_check",
]
WIDTHS = {
    "id": 8, "product_id": 12, "product_url": 30, "product_name": 28,
    "product_type": 15, "certified_function": 40, "disclaimer": 13,
    "tile_text": 50, "context_before": 30,
    "sentence": 46, "context_after": 30, "label": 16, "confidence": 11,
    "note": 28, "labeler": 10, "is_cross_check": 13,
}

# 건강기능식품은 "영양·기능정보"란에 인정받은 기능성을 의무 표기한다.
# 광고 문구가 이 범위 안인지 대조하는 것이 3호·4호 판정의 핵심이라,
# 상세페이지를 따로 열지 않고 시트에서 바로 볼 수 있게 뽑아 넣는다.
CERTIFIED = re.compile(
    r"도움을?\s*줄\s*수\s*있(음|습니다)|기능성\s*내용|영양[·ㆍ]\s*기능정보"
)


PRODUCT_TYPES = ["일반식품", "건강기능식품", "불명"]
CONFIDENCE = ["확실", "애매"]
MAX_ROW = 400  # 검증·수식이 미리 깔리는 범위


def col(name: str) -> str:
    """컬럼명 → 엑셀 열 문자."""
    return get_column_letter(COLUMNS.index(name) + 1)


def normalize(s: str) -> str:
    """중복 판정용 정규화."""
    return re.sub(r"[\s\W_]+", "", s)


# 시행령 별표1 1호 라목 단서 2) — 건강기능식품이 질병정보를 쓰면서
# "제품과 직접적인 관련이 없습니다"를 병기하면 라목에서 제외된다.
DISCLAIMER = re.compile(
    r"제품과\s*(직접적인\s*)?관련\s*이?\s*없|관련\s*없는\s*건강\s*정보"
    r"|제품이\s*아닌\s*제품\s*원료|원료에\s*관한\s*설명"
)


def has_disclaimer(sentences: list[str]) -> bool:
    """질병정보 면책 병기가 상세페이지에 있는지."""
    return any(DISCLAIMER.search(t) for t in sentences)


def extract_certified(sentences: list[str], limit: int = 4) -> str:
    """OCR 문장에서 인정 기능성 표기를 뽑는다. 없으면 빈 문자열."""
    out, seen = [], set()
    for text in sentences:
        if not CERTIFIED.search(text):
            continue
        t = re.sub(r"^[·ㆍ\-\s]+", "", text).strip()
        # 표제어만 있는 조각("영양·기능정보")은 정보가 없으니 뺀다.
        if len(t) < 10 or normalize(t) in seen:
            continue
        seen.add(normalize(t))
        out.append(t if len(t) <= 70 else t[:70] + "…")
        if len(out) >= limit:
            break
    return "\n".join(out)


def is_korean(s: str, min_syllables: int = 3, min_ratio: float = 0.3) -> bool:
    """한국어 문장인지 본다.

    해외 직구 상품 상세페이지에는 기계번역된 영문이 섞여 들어온다.
    학생 모델이 한국어 문장 분류기라 영문 문장은 홀드아웃에서 뺀다.
    """
    hangul = len(re.findall(r"[가-힣]", s))
    letters = len(re.findall(r"[가-힣A-Za-z]", s))
    return hangul >= min_syllables and (not letters or hangul / letters >= min_ratio)


# 디자인상 줄이 나뉜 문구가 조각으로 끊겨 나오는 경우를 거른다.
# 연결어미·조사로 끝나면 뒷말이 잘린 것이라 문장만 보고는 판정이 불가능하다.
TRUNCATED_END = re.compile(
    r"(되지|하지|않고|으로|에서|보다|처럼|까지|부터|이라|라고|해서|하고|이며|하며"
    r"|인지|는지|은지|지만|면서|려면|는데|은데|어야|아야|다가)$"
)


def is_fragment(s: str) -> bool:
    """판정 불가능한 문장 조각인지 본다."""
    t = s.strip().rstrip("!?.…~ ")
    return t.endswith((",", "·", "-", "및")) or bool(TRUNCATED_END.search(t))


# 반려동물 식품은 사료관리법 소관이라 식품표시광고법 판정 대상이 아니다.
# 크롤 검색어("관절"·"수면" 등)에 반려동물 상품이 딸려 온다.
PET = re.compile(r"강아지|애견|반려견|반려묘|반려동물|고양이|노견|노묘|퍼피|펫|사료|수의사")


def is_pet(*texts: str) -> bool:
    """반려동물 상품·문장인지 본다."""
    return any(PET.search(t or "") for t in texts)


def load_ocr() -> dict[str, dict]:
    """OCR 결과를 모든 샤드 파일에서 읽는다."""
    out = {}
    for path in sorted(Path("data").glob("ocr_sentences*.jsonl")):
        for line in open(path):
            if line.strip():
                rec = json.loads(line)
                prev = out.get(rec["product_id"])
                if prev and prev["sentences"] and not rec["sentences"]:
                    continue
                out[rec["product_id"]] = rec
    return out


def load_urls() -> dict[str, str]:
    """매니페스트에서 상품코드 → 상세 URL."""
    urls = {}
    for f in glob.glob("11st_output/11st_details_*.json"):
        for p in json.load(open(f)).get("products", []):
            if p.get("detail_url"):
                urls[p["product_code"]] = p["detail_url"]
    return urls


def load_prescreen() -> dict[str, dict]:
    """선별 결과를 모두 읽어 상품별로 합친다.

    모델을 바꿔 2차 재판정한 결과가 있으면 함께 병합한다. 같은 문장을 두 모델이
    다르게 봤다면 **더 무거운 쪽**을 택한다 — 미탐(놓친 위반)이 오탐보다 비싸고,
    최종 판단은 어차피 사람이 하므로 후보를 넓게 잡는 편이 낫다.
    """
    rank = {label: i for i, label in enumerate(SEVERITY)}
    merged: dict[str, dict] = {}

    for path in sorted(Path("data").glob(PRESCREEN_GLOB)):
        for line in open(path):
            if not line.strip():
                continue
            rec = json.loads(line)
            cur = merged.get(rec["product_id"])
            if cur is None:
                merged[rec["product_id"]] = {
                    **rec,
                    "sentences": {s["order"]: s for s in rec["sentences"]},
                }
                continue

            # product_type은 '불명'을 확정 판정으로 덮어쓴다.
            if cur["product_type"] == "불명" and rec["product_type"] != "불명":
                cur["product_type"] = rec["product_type"]
                cur["product_type_evidence"] = rec["product_type_evidence"]

            for s in rec["sentences"]:
                prev = cur["sentences"].get(s["order"])
                if prev is None or rank[s["hint"]] < rank[prev["hint"]]:
                    cur["sentences"][s["order"]] = s

    for rec in merged.values():
        rec["sentences"] = [rec["sentences"][k] for k in sorted(rec["sentences"])]
    return merged


def build_pool() -> list[dict]:
    """OCR 원문 + 선별 결과를 합쳐 후보 문장 풀을 만든다.

    문맥(context_before/after)은 **선별 전 원본 순서**에서 가져온다 —
    걸러진 문장도 라벨러에겐 판단 재료이므로 그대로 붙인다.
    """
    ocr = load_ocr()
    urls = load_urls()
    pool, seen = [], set()

    skipped_pet = 0
    for pre in load_prescreen().values():
        src = ocr.get(pre["product_id"])
        if not src:
            continue
        texts = [s["text"] for s in src["sentences"]]

        # 상품명 또는 상세 전문에 반려동물 표현이 있으면 상품 통째로 제외한다.
        if is_pet(pre["product_name"], " ".join(texts[:80])):
            skipped_pet += 1
            continue

        # 같은 타일 블록 전체 텍스트 — 논문 인용정보·후기 본문이 같은 타일에 있으면
        # 라벨러가 페이지를 안 열고도 판정할 수 있다(모델의 타일 문맥과 같은 단위).
        tile_map = defaultdict(list)
        for s in src["sentences"]:
            tile_map[s["tile"]].append(s["text"])

        for s in pre["sentences"]:
            text = s["text"].strip()
            key = normalize(text)
            if (len(key) < 8 or key in seen
                    or not is_korean(text) or is_fragment(text)
                    or key in normalize(pre["product_name"])):
                continue  # 짧은 조각·중복·영문·잘린 문장·상품명 반복(제품명 표기)
            seen.add(key)
            i = s["order"]
            block = " / ".join(tile_map.get(s.get("tile", ""), []))
            pool.append({
                "product_id": pre["product_id"],
                "product_url": urls.get(pre["product_id"], ""),
                "product_name": pre["product_name"],
                "product_type": pre["product_type"],
                "tile_text": block if len(block) <= 600 else block[:600] + "…",
                "context_before": " / ".join(texts[max(0, i - 2):i]),
                "sentence": text,
                "context_after": " / ".join(texts[i + 1:i + 3]),
                "hint": s["hint"],
            })

    if skipped_pet:
        print(f"[제외] 반려동물 상품 {skipped_pet}개")
    return pool


def allocate(pool: list[dict], total: int, per_product: int) -> list[dict]:
    """유형 힌트 기준으로 층화 추출한다.

    힌트는 부정확하므로 쿼터보다 넉넉히 뽑는다. 부족한 유형의 몫은
    남는 유형으로 넘겨 총량을 채운다.

    **상품당 문장 수에 상한을 둔다.** 상한이 없으면 상세페이지가 긴 상품 몇 개가
    홀드아웃을 점령해, 그 상품의 문체·표현만 평가하는 꼴이 된다.
    """
    by_hint = defaultdict(list)
    for row in pool:
        by_hint[row["hint"]].append(row)
    for rows in by_hint.values():
        random.shuffle(rows)

    scale = total / sum(QUOTA.values())
    want = {k: round(v * scale) for k, v in QUOTA.items()}
    used = defaultdict(int)

    def take_from(rows: list[dict], n: int, cap: int) -> list[dict]:
        """상품당 cap을 지키며 n개까지 뽑고, 뽑은 것은 rows에서 제거한다."""
        out, rest = [], []
        for row in rows:
            if len(out) < n and used[row["product_id"]] < cap:
                used[row["product_id"]] += 1
                out.append(row)
            else:
                rest.append(row)
        rows[:] = rest
        return out

    # 희소한 유형부터 배정한다. 흔한 유형(합법)이 먼저 가져가면 상품당 상한을
    # 다 써버려, 정작 부족한 위반 유형이 뽑을 상품이 남지 않는다.
    picked, shortfall = [], 0
    for label in sorted(want, key=lambda k: len(by_hint[k])):
        n = want[label]
        # 위반 문구는 원래 일부 판매자에게 몰려 있으므로 상한을 넉넉히 준다.
        # 합법은 풍부하니 좁게 잡아 상품을 넓게 퍼뜨린다.
        cap = per_product // 2 if label == "합법" else per_product
        got = take_from(by_hint[label], n, cap)
        picked += got
        shortfall += n - len(got)
        if len(got) < n:
            print(f"  [부족] {label}: {len(got)}/{n} "
                  f"(풀 {len(got) + len(by_hint[label])}개, 상품당 {cap}개 상한)")

    # 남은 몫은 여유 있는 유형에서 채운다(대상외 포함).
    leftovers = [r for rows in by_hint.values() for r in rows]
    random.shuffle(leftovers)
    picked += take_from(leftovers, shortfall, per_product)
    return picked


def split_sheets(picked: list[dict], shared: int) -> tuple[list, list]:
    """공통 블록 + 각자 고유분으로 나눈다.

    두 사람이 비슷한 난이도를 받도록 유형별로 번갈아 배분한다.
    공통 블록도 유형이 골고루 섞이게 뽑는다 — 상단 20행은 사전 정렬(기준서 §5.1),
    전체 30행은 교차검증(§5.3)에 쓴다.
    """
    by_hint = defaultdict(list)
    for row in picked:
        by_hint[row["hint"]].append(row)
    hints = sorted(by_hint)

    # 공통 블록은 두 사람이 같이 보고 기준을 맞추는 곳이다. 한 상품 문장이 겹치면
    # 같은 맥락을 반복해 보게 되어 기준 검증력이 떨어지므로 상품을 전부 다르게 뽑는다.
    common, used_products = [], set()
    for allow_dup in (False, True):
        while len(common) < shared:
            added = False
            for h in hints:
                if len(common) >= shared:
                    break
                for i, row in enumerate(by_hint[h]):
                    if allow_dup or row["product_id"] not in used_products:
                        used_products.add(row["product_id"])
                        common.append(by_hint[h].pop(i))
                        added = True
                        break
            if not added:
                break

    # 유형별로 번갈아 나눠 A/B의 유형 분포를 맞춘다.
    only_a, only_b = [], []
    for h in hints:
        for i, row in enumerate(by_hint[h]):
            (only_a if i % 2 == 0 else only_b).append(row)

    random.shuffle(only_a)
    random.shuffle(only_b)
    return common + only_a, common + only_b


def _write_guide(ws) -> None:
    """라벨러용 안내 시트."""
    rows = [
        ["라벨링 작업 안내"],
        [],
        ["barum — 허위·과대광고 평가 홀드아웃 구축"],
        [],
        ["무엇을 하는 건가"],
        ["", "우리 AI가 광고 문구를 보고 위반 여부를 판정합니다. 그 AI가 얼마나 정확한지 재려면"],
        ["", "사람이 찍은 정답이 필요합니다. AI가 만든 답으로 AI를 채점하면 같은 실수를 못 잡습니다."],
        ["", "이 파일에 채워 넣는 label이 그 정답지가 됩니다."],
        [],
        ["시작 전 필수 — 30분"],
        ["", "바로 본 작업에 들어가지 마세요. is_cross_check=Y 인 상단 20행을 각자 먼저 라벨링하고"],
        ["", "결과를 맞춰 봅니다."],
        ["", "· 16개 이상 일치 → 본 작업 시작"],
        ["", "· 미만 → 어긋난 부분을 논의해 기준을 맞추고 다시 20개"],
        ["", "이 30분을 아끼면 나중에 200건을 다시 봐야 합니다."],
        [],
        ["작업 방법"],
        ["", "1. sentence 열의 문장을 봅니다. 앞뒤 문맥이 필요하면 context_before/after를 참고합니다."],
        ["", "2. label 열에서 드롭다운으로 하나를 고릅니다."],
        ["", "3. 확신이 없으면 confidence=애매 로 두고 note에 무엇이 걸리는지 적습니다."],
        ["", "4. labeler 열에 본인 이름을 적습니다."],
        ["", "억지로 찍지 마세요. 애매 표시된 건은 나중에 모여서 함께 결정합니다."],
        [],
        ["가장 중요한 규칙 — product_type을 먼저 보세요"],
        ["", "같은 문장이라도 제품 유형에 따라 결론이 반대가 됩니다."],
        ["", "\"체지방 감소에 도움\"  →  일반식품이면 3호 위반 / 건강기능식품이면 합법"],
        ["", "\"면역력 증진에 도움\"  →  일반식품이면 3호 위반 / 건강기능식품이면 합법"],
        ["", "product_type이 '불명'이면 product_url을 열어 확인하고 값을 고쳐 주세요."],
        [],
        ["certified_function 열 — 건강기능식품 판정의 기준자"],
        ["", "그 제품이 식약처에서 인정받은 기능성 문구입니다. 상세페이지의 '영양·기능정보'에서 뽑았습니다."],
        ["", "건강기능식품이라도 인정받은 범위 안에서만 말할 수 있습니다."],
        ["", "· 광고 문구가 이 범위 안  →  합법"],
        ["", "· 범위 밖의 신체 작용     →  4호"],
        ["", "예) 인정문구가 '전립선 건강의 유지에 도움'인데 광고가 '잔뇨감이 있으신 분' → 범위 밖"],
        ["", "이 열이 비어 있으면 인정 기능성이 없는 제품입니다(대개 일반식품)."],
        [],
        ["tile_text 열 — 같은 화면(타일)의 전체 문구"],
        ["", "sentence가 나온 타일 이미지의 모든 문구를 모아둔 것입니다."],
        ["", "논문 인용정보(저자·연도)나 후기 본문이 같은 화면에 있으면 여기서 보입니다."],
        ["", "예) sentence가 '12주 임상 결과'인데 tile_text에 '저자·논문명·2004'가 있으면 인용정보 확인 가능."],
        ["", "sentence만으로 애매하면 반드시 tile_text를 먼저 확인하세요. 그래도 정보가 없으면 그때 애매."],
        [],
        ["\"사실이면 괜찮은가?\" — 아닙니다"],
        ["", "진위를 따지는 건 4호뿐입니다. 1·2·3·5호는 '인식할 우려'가 요건이라"],
        ["", "내용이 사실이어도 성립합니다."],
        ["", "\"비타민C는 항염 효과가 있습니다\"는 사실이어도 위반입니다."],
        [],
        ["수치·특허·논문이 나와도 애매로 빼지 마세요"],
        ["", "'120배'가 진짜인지 확인하려 하지 마세요. 표현만 보고 확정합니다."],
        ["", "'체내 흡수율 120배' → 흡수율이라는 신체 작용을 말했으므로 4호. 수치의 진위와 무관."],
        ["", "note에 '수치 주장' 또는 '특허 주장'이라고만 적어 주세요."],
        ["", "애매는 어느 호인지 갈릴 때 쓰는 것이지, 사실 확인이 안 될 때 쓰는 게 아닙니다."],
        [],
        ["한 문장에 여러 유형이 보이면"],
        ["", "위해등급이 높은 것 하나만 고릅니다.  2호 → 1호 → 3호 → 4호 → 5호"],
        ["", "note에 '1호+5호 복합'처럼 적어 주세요."],
        [],
        ["6~10호로 보이면"],
        ["", "억지로 6개 중 하나에 밀어넣지 말고 label=대상외 로 두고 note에 몇 호로 보이는지 적습니다."],
        [],
        ["발견한 건강기능식품 인정 문구는 note에 적어 주세요 — 나중에 기준 문서의 초안이 됩니다."],
        [],
        ["상세 판단 기준은 [판정기준] 시트와 라벨링_기준서.md 를 보세요."],
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)
    for i, r in enumerate(rows, start=1):
        if r and len(r) == 1 and i > 1:
            ws.cell(row=i, column=1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 96


def _write_criteria(ws) -> None:
    """판정 기준 요약 시트."""
    rows = [
        ["라벨링 판정 기준 요약"],
        ["근거: 「식품 등의 표시·광고에 관한 법률」 제8조 제1항 / 시행령 제3조 제1항 및 별표 1"],
        [],
        ["1. 라벨 6개 + 대상외"],
        ["라벨", "무엇을 잡나", "실제 적발 문구", "위해등급", "법정형"],
        ["합법", "위반 아님. 건강기능식품의 인정 기능성 표현 포함", "—", "—", "—"],
        ["1호_질병표방", "질병을 예방·치료 / 질병의 징후·증상에 효과 / 질병정보를 활용해 연관 암시",
         "암 예방, 변비 개선, 관절염 통증 개선, 역류성식도염 개선", "최고", "10년 / 1억"],
        ["2호_의약품오인", "의약품에만 쓰는 명칭 / 의약품에 포함된다 / 의약품을 대체 / 의약품 효능을 증대",
         "다이어트약, 간장약, 키크는 약, 위고비, GLP-1", "최고", "10년 / 1억"],
        ["3호_건기식오인", "일반식품인데 건강기능식품법상 '기능성'이 있는 것처럼 표현",
         "영양제, 면역력 강화, 체지방 감소", "高", "10년 / 1억"],
        ["4호_거짓과장", "신고사항과 다름 / 사실과 다름 / 신체조직의 기능·작용·효과를 표현 / 수상·인증·특허를 사실과 다르게",
         "붓기 제거, 피부 탄력, 긴장 완화, 특허받은(특허 없음)", "中", "5년 / 5천"],
        ["5호_소비자기만", "원재료 효능을 제품 효능으로 / 체험기·감사장 / 의사·약사 등 전문가 보증 / 공인 안 된 제조방법 인용",
         "약사 추천, 3개월 -10kg 후기, 알부민 효능·효과", "中", "5년 / 5천"],
        ["대상외", "6~10호(비방, 부당비교, 사행심, 유사포장, 자율심의 미필)", "—", "—", "—"],
        [],
        ["2. 헷갈리는 경계"],
        ["문구", "라벨", "이유"],
        ["관절염 개선", "1호", "관절염은 질병명"],
        ["관절 건강에 도움 (일반식품)", "3호", "질병이 아니라 기능성 표현"],
        ["관절 건강에 도움 (건강기능식품)", "합법", "인정 기능성"],
        ["체지방 감소에 도움 (일반식품)", "3호", "건기식 인정 기능성 문구를 그대로 사용"],
        ["다리 붓기 빼주는", "4호", "인정 기능성 목록에 없는 신체조직 작용"],
        ["피부 탄력 개선", "4호", "별표1 4호 다목 — 신체조직의 기능·작용"],
        ["특허받은 성분 (특허 없음)", "4호", "4호 라목 — 특허 관련 사실과 다름"],
        ["○○ 성분은 항염 효과가 있습니다", "5호", "5호 나목 — 원재료 효능을 제품 효능으로"],
        ["약사가 추천하는", "5호", "5호 라목 — 전문가 보증"],
        ["3개월만에 -10kg 성공! (후기)", "5호", "5호 다목 — 체험기"],
        [],
        [],
        ["3. 판정 순서 — product_type과 certified_function을 먼저 봅니다"],
        ["제품 유형", "표현", "라벨", "", ""],
        ["건강기능식품", "certified_function 범위 안", "합법", "", ""],
        ["건강기능식품", "인정받지 않은 기능성·신체작용", "4호", "인정 범위 월권", ""],
        ["일반식품", "건기식 인정 기능성 표현", "3호", "\"체지방 감소에 도움\"", ""],
        ["일반식품", "인정 목록에도 없는 신체작용", "4호", "\"붓기 제거\", \"흡수율 120배\"", ""],
        ["", "3호는 '건기식이 아닌 것을 건기식으로 오인'시키는 조항이라 건기식 제품에는 적용되지 않습니다.", "", "", ""],
        [],
        ["4. 증상 나열 — 1호 라목"],
        ["", "라목의 '질병정보'에는 질병의 징후·증상이 포함됩니다. 효과 동사도 질병명도 없어도 성립합니다.", "", "", ""],
        ["문구", "라벨", "이유", "", ""],
        ["잔뇨감이 있으신 분 / 밤낮으로 화장실을 자주", "1호", "전립선비대증의 특징적 증상 = 질병정보", "", ""],
        ["소변줄기가 가늘고 힘이 없어진다", "1호", "같음", "", ""],
        ["피곤하다 / 입맛이 없다 / 생각이 많아진다", "합법", "특정 질병의 진단 단서가 아닌 일반 컨디션", "", ""],
        ["", "갈림길: 그 증상이 특정 질병의 진단 단서로 쓰이는가.", "", "", ""],
        [],
        ["5. 원재료가 주어인 문장"],
        ["원재료 효능의 내용", "라벨", "예시", "", ""],
        ["질병 관련", "1호", "\"비타민C는 항암 효과\", \"○○는 항염 효과\"", "", ""],
        ["건기식 인정 기능성 (제품은 일반식품)", "3호", "\"○○는 체지방 감소에 도움\"", "", ""],
        ["인정 목록에 없는 신체 작용", "5호", "\"생체이용률 3.04배\", \"색소 침착 개선 논문\"", "", ""],
        ["성분 함유 사실만 진술", "합법", "\"우리 제품은 비타민C를 함유\"", "", ""],
        ["", "라벨은 효능을 주장한 문장에 답니다. 성분 함유 문장은 합법으로 두고 note에 '앞 문장과 결합 시 오인'.", "", "", ""],
        ["", "면책 문구('원재료에 대한 설명에 한합니다')는 그 자체로는 합법이지만, 앞의 효능 서술을 면책해 주지 않습니다.", "", "", ""],
        [],
        ["6. 4호는 note에 하위 목을 적어 주세요"],
        ["", "다목 — 신체조직의 기능·작용 (\"붓기 제거\", \"피부 탄력\"). 진위 확인 불필요", "", "", ""],
        ["", "라목 — 수상·인증·보증·특허 (\"특허받은 성분\", \"1위\"). 진위 확인 필요", "", "", ""],
        ["", "나중에 '진위 확인이 필요한 건'과 아닌 건을 분리해 평가하기 위한 표시입니다.", "", "", ""],
        [],
        ["갈림길 정리"],
        ["", "질병명·질병의 징후·증상이 나오면 1호"],
        ["", "건기식 인정 기능성 표현을 일반식품이 쓰면 3호, 인정 목록에도 없는 신체작용이면 4호"],
        ["", "사실관계 자체가 틀렸으면 4호, 사실은 맞는데 오인하게 만드는 방식이면 5호"],
        [],
        ["3. 실제 적발 분포 (식약처 온라인 점검)"],
        ["유형", "2026.06 (225건)", "2025.06 (236건)"],
        ["3호 건기식 오인", "46.2%", "41.1%"],
        ["1호 질병표방", "37.3%", "31.4%"],
        ["5호 소비자 기만", "8.5%", "9.7%"],
        ["2호 의약품 오인", "4.4%", "3.4%"],
        ["4호 거짓·과장", "3.6%", "14.0%"],
        [],
        ["가장 적게 잡히는 2호가 가장 위험한 유형입니다. 드물어서 놓치기 쉽고, 놓치면 제일 위험합니다."],
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)
    for i, r in enumerate(rows, start=1):
        if r and r[0] and str(r[0])[0].isdigit() and str(r[0])[1:2] == ".":
            ws.cell(row=i, column=1).font = Font(bold=True)
    for c, w in zip("ABCDE", (34, 52, 46, 12, 14)):
        ws.column_dimensions[c].width = w
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_progress(ws, n_rows: int) -> None:
    """유형별 확보 현황 — 라벨링 시트를 COUNTIF로 집계한다."""
    lc, cc, mc = col("label"), col("confidence"), col("is_cross_check")
    rng = f"라벨링!${lc}$2:${lc}${MAX_ROW}"

    ws.append(["유형별 확보 현황"])
    ws.append(["라벨", "목표", "현재", "달성률", "비고"])
    notes = {
        "합법": "쉬움",
        "1호_질병표방": "쉬움 — 적발 분포 31~37%",
        "2호_의약품오인": "어려움 — 분포 3~4%. 의도적 검색 필요",
        "3호_건기식오인": "쉬움 — 적발 분포 1위 41~46%",
        "4호_거짓과장": "보통",
        "5호_소비자기만": "보통",
        "대상외": "6~10호 해당. 목표 없음",
    }
    order = ["합법", "1호_질병표방", "2호_의약품오인",
             "3호_건기식오인", "4호_거짓과장", "5호_소비자기만", "대상외"]
    for i, label in enumerate(order, start=3):
        ws.append([
            label, QUOTA.get(label, 0), f'=COUNTIF({rng},A{i})',
            f'=IF(B{i}=0,"-",C{i}/B{i})', notes[label],
        ])
    last = 2 + len(order)
    ws.append(["합계", f"=SUM(B3:B{last})", f"=SUM(C3:C{last})", "",
               "목표 총량 200~300문장 (두 시트 합산 기준)"])
    ws.append([])
    ws.append(["품질 지표"])
    ws.append(["작성 완료", f'=COUNTA(라벨링!${lc}$2:${lc}${MAX_ROW})',
               "", "", f"이 시트 전체 {n_rows}행"])
    ws.append(["애매 표시", f'=COUNTIF(라벨링!${cc}$2:${cc}${MAX_ROW},"애매")',
               "", "", "모여서 함께 결정할 건"])
    ws.append(["교차검증 대상", f'=COUNTIF(라벨링!${mc}$2:${mc}${MAX_ROW},"Y")',
               "", "", "두 사람이 같이 라벨링하는 공통 문장"])
    ws.append([])
    ws.append(["※ 달성률은 두 시트를 합쳐서 봐야 합니다. 이 시트 숫자는 본인 몫입니다."])

    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"].font = Font(bold=True)
    ws[f"A{last + 3}"].font = Font(bold=True)
    for c, w in zip("ABCDE", (18, 10, 10, 10, 40)):
        ws.column_dimensions[c].width = w
    for i in range(3, last + 1):
        ws[f"D{i}"].number_format = "0%"


def write_xlsx(rows: list[dict], path: Path, sheet_tag: str) -> None:
    """배포용 xlsx를 쓴다. label·confidence·product_type은 드롭다운으로 고정한다."""
    wb = Workbook()

    _write_guide(wb.active)
    wb.active.title = "시작하기"

    ws = wb.create_sheet("라벨링")
    ws.append(COLUMNS)
    header_fill = PatternFill("solid", fgColor="DDE7F0")
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # 라벨러가 채우는 칸은 비워서 배포한다(정답 유출 방지).
    blank = {"label", "confidence", "note", "labeler"}
    for row in rows:
        ws.append(["" if c in blank else row.get(c, "") for c in COLUMNS])

    for name, options in (
        ("label", LABELS),
        ("confidence", CONFIDENCE),
        ("product_type", PRODUCT_TYPES),
        ("is_cross_check", ["Y", "N"]),
        ("disclaimer", ["있음", "없음"]),
    ):
        c = col(name)
        dv = DataValidation(
            type="list", formula1=f'"{",".join(options)}"', allow_blank=True,
            showDropDown=False, errorTitle="허용되지 않는 값",
            error="드롭다운 목록에서 골라 주세요.",
        )
        ws.add_data_validation(dv)
        dv.add(f"{c}2:{c}{MAX_ROW}")

    for i, name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS[name]
    wrap_cols = [COLUMNS.index(n) + 1 for n in
                 ("product_name", "certified_function", "tile_text",
                  "context_before", "sentence", "context_after")]
    for r in range(2, len(rows) + 2):
        for i in wrap_cols:
            ws.cell(row=r, column=i).alignment = Alignment(
                wrap_text=True, vertical="top")

    # 교차검증 공통 블록을 색으로 구분(상단 20행 = 사전 정렬용).
    cross_fill = PatternFill("solid", fgColor="FFF2CC")
    align_fill = PatternFill("solid", fgColor="FCE4D6")
    for r in range(2, len(rows) + 2):
        if ws.cell(row=r, column=COLUMNS.index("is_cross_check") + 1).value == "Y":
            fill = align_fill if r <= 21 else cross_fill
            for i in range(1, len(COLUMNS) + 1):
                ws.cell(row=r, column=i).fill = fill

    ws.freeze_panes = "B2"

    _write_progress(wb.create_sheet("진행현황"), len(rows))
    _write_criteria(wb.create_sheet("판정기준"))
    wb.save(path)


def rebuild_from_master(master: Path) -> None:
    """확정된 master의 행 순서를 그대로 두고 시트만 다시 만든다.

    라벨링이 진행 중일 때 컬럼을 추가하려면 행이 흔들리면 안 된다.
    """
    ocr = load_ocr()
    rows = [json.loads(l) for l in open(master)]

    certified, disclaimer = {}, {}
    for pid in {r["product_id"] for r in rows}:
        src = ocr.get(pid)
        texts = [s["text"] for s in src["sentences"]] if src else []
        certified[pid] = extract_certified(texts)
        disclaimer[pid] = "있음" if has_disclaimer(texts) else "없음"

    for r in rows:
        r["certified_function"] = certified.get(r["product_id"], "")
        r["disclaimer"] = disclaimer.get(r["product_id"], "없음")

    for tag in ("A", "B"):
        sheet = [r for r in rows if r["sheet"] == tag]
        write_xlsx(sheet, Path(f"data/holdout_{tag}.xlsx"), tag)
        n = sum(1 for r in sheet if r["certified_function"])
        d = sum(1 for r in sheet if r["disclaimer"] == "있음")
        print(f"{tag} 시트: {len(sheet)}행 (인정문구 {n}행 / 면책병기 {d}행)")

    with open(master, "w") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    have = sum(1 for v in certified.values() if v)
    print(f"상품 {len(certified)}개 중 {have}개에서 인정 기능성 문구 추출")

    _run_validation({t: [r for r in rows if r["sheet"] == t] for t in ("A", "B")})


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--total", type=int, default=300, help="유니크 문장 총량")
    ap.add_argument("--shared", type=int, default=30, help="두 시트 공통 문장 수")
    ap.add_argument("--seed", type=int, default=20260805)
    ap.add_argument("--per-product", type=int, default=6,
                    help="한 상품에서 뽑을 문장 수 상한")
    ap.add_argument("--exclude", nargs="*", default=[],
                    help="이 master jsonl들에 쓰인 문장은 제외(중복 방지)")
    ap.add_argument("--out-prefix", default="holdout",
                    help="출력 파일 접두사 (data/{prefix}_A.xlsx 등)")
    ap.add_argument("--from-master",
                    help="기존 master jsonl의 행을 그대로 쓰고 시트만 다시 만든다 "
                         "(라벨링 중인 시트에 컬럼만 추가할 때)")
    args = ap.parse_args()

    random.seed(args.seed)

    if args.from_master:
        rebuild_from_master(Path(args.from_master))
        return

    pool = build_pool()
    if args.exclude:
        used = set()
        for f in args.exclude:
            for line in open(f):
                if line.strip():
                    used.add(normalize(json.loads(line)["sentence"]))
        before = len(pool)
        pool = [r for r in pool if normalize(r["sentence"]) not in used]
        print(f"[제외] 이전 사용 문장 {len(used)}개 → 풀 {before}→{len(pool)}")
    print(f"후보 문장 풀: {len(pool)}개")
    dist = defaultdict(int)
    for r in pool:
        dist[r["hint"]] += 1
    for k, v in sorted(dist.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")

    print(f"\n층화 추출 (목표 {args.total}개)")
    picked = allocate(pool, args.total, args.per_product)
    products = {r["product_id"] for r in picked}
    print(f"→ {len(picked)}개 선택 (상품 {len(products)}개)")

    sheet_a, sheet_b = split_sheets(picked, args.shared)

    # id는 공통(S)/고유(A/B)를 구분해 붙인다. 집계 때 공통 블록을 바로 골라낼 수 있다.
    for sheet, tag in ((sheet_a, "A"), (sheet_b, "B")):
        for i, row in enumerate(sheet):
            shared_row = i < args.shared
            row["id"] = f"S{i + 1:03d}" if shared_row else f"{tag}{i - args.shared + 1:03d}"
            row["is_cross_check"] = "Y" if shared_row else "N"

    px = args.out_prefix
    write_xlsx(sheet_a, Path(f"data/{px}_A.xlsx"), "A")
    write_xlsx(sheet_b, Path(f"data/{px}_B.xlsx"), "B")

    with open(f"data/{px}_master.jsonl", "w") as f:
        for sheet, tag in ((sheet_a, "A"), (sheet_b, "B")):
            for row in sheet:
                f.write(json.dumps({**row, "sheet": tag}, ensure_ascii=False) + "\n")

    for tag, sheet in (("A", sheet_a), ("B", sheet_b)):
        d = defaultdict(int)
        for r in sheet:
            d[r["hint"]] += 1
        mix = " ".join(f"{k.split('_')[0]}:{v}" for k, v in sorted(d.items()))
        print(f"\n{tag} 시트: {len(sheet)}행 (공통 {args.shared} + 고유 "
              f"{len(sheet) - args.shared})\n  힌트 분포 {mix}")

    print(f"\n→ data/{px}_A.xlsx, data/{px}_B.xlsx")
    print(f"→ data/{px}_master.jsonl (힌트 포함, 배포 금지)")

    _run_validation({"A": sheet_a, "B": sheet_b})


def _run_validation(sheets: dict) -> None:
    """빌드 끝에 기계 검수를 돌린다. ERROR가 있으면 크게 알린다.

    지연 import — validate_holdout이 이 모듈을 import하므로 순환을 피한다.
    """
    from validate_holdout import validate_rows

    print("\n=== 기계 검수 ===")
    rep = validate_rows(sheets)
    rep.print()
    if rep.errors:
        print("‼️  ERROR가 있습니다. 배포 전에 반드시 확인하세요.")


if __name__ == "__main__":
    main()
