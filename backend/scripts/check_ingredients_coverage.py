# -*- coding: utf-8 -*-
"""전성분 커버리지 확인: 53장 OCR 캐시(label_worksheet_combined.xlsx)에서 몇 장이
OCR 텍스트만으로 전성분 추출 가능한지 센다. API 호출 없음(캐시된 OCR 결과 재사용).

실행(backend/에서):
  python scripts/check_ingredients_coverage.py

출력:
  11st_probe_cosmetic/read_test/ingredients_map.json  (nn -> ingredients|null, source)
  11st_probe_cosmetic/read_test/_run_log_ingredients_coverage.txt
"""
import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, "src")
from barum.preprocess.ingredients import extract_ingredients_block  # noqa: E402

_READ_TEST = Path("11st_probe_cosmetic/read_test")
_LABEL_XLSX = _READ_TEST / "label_worksheet_combined.xlsx"
_ANSWER_KEY = _READ_TEST / "_combined_answer_key.json"
_OUT_MAP = _READ_TEST / "ingredients_map.json"
_OUT_LOG = _READ_TEST / "_run_log_ingredients_coverage.txt"


def load_sentences_by_image() -> tuple[dict[str, list[dict]], dict[str, str]]:
    """label_worksheet_combined.xlsx에서 이미지별 OCR 문장·상품코드를 읽는다."""
    wb = openpyxl.load_workbook(_LABEL_XLSX)
    ws = wb["라벨링"]
    by_img: dict[str, list[dict]] = {}
    codes: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        nn = str(ws.cell(r, 1).value or "").strip()
        code = str(ws.cell(r, 2).value or "").strip()
        sent = str(ws.cell(r, 4).value or "").strip()
        if not nn:
            continue
        by_img.setdefault(nn, []).append({"text": sent})
        codes.setdefault(nn, code)
    return by_img, codes


def main() -> None:
    all_nn = [item["nn"] for item in json.loads(_ANSWER_KEY.read_text(encoding="utf-8"))]
    by_img, codes = load_sentences_by_image()

    result: dict[str, dict] = {}
    hit, miss, no_ocr = [], [], []
    for nn in all_nn:
        sentences = by_img.get(nn)
        if not sentences:
            no_ocr.append(nn)
            result[nn] = {"product_code": codes.get(nn), "ingredients": None, "source": "no_ocr_cache"}
            continue
        block = extract_ingredients_block(sentences)
        if block:
            hit.append(nn)
            result[nn] = {"product_code": codes.get(nn), "ingredients": block, "source": "ocr_text"}
        else:
            miss.append(nn)
            result[nn] = {"product_code": codes.get(nn), "ingredients": None, "source": None}

    _OUT_MAP.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        f"전성분 커버리지: {len(hit)}/{len(all_nn)} ({len(hit) / len(all_nn):.1%})",
        f"OCR 텍스트로 추출됨({len(hit)}): {', '.join(hit)}",
        f"추출 실패, 보완 크롤 필요({len(miss)}): {', '.join(miss)}",
    ]
    if no_ocr:
        lines.append(f"OCR 캐시 자체가 없음({len(no_ocr)}): {', '.join(no_ocr)}")
    log_text = "\n".join(lines)
    _OUT_LOG.write_text(log_text + "\n", encoding="utf-8")
    print(log_text)
    print(f"\n저장: {_OUT_MAP}")


if __name__ == "__main__":
    main()
