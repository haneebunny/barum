"""홀드아웃 시트 기계 검수 — LLM 없이 결정적으로 잡히는 결함만 잡는다.

단독 실행:  venv/bin/python scripts/validate_holdout.py
빌드 연동:  build_holdout 에서 validate_rows() 를 호출한다.

ERROR 가 하나라도 있으면 종료코드 1 (배포 차단용). WARN 은 보고만 한다.
검사 로직은 순수 함수라 tests/test_validate.py 로 유닛테스트한다.
"""

import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from build_holdout import (  # noqa: E402
    COLUMNS, LABELS, QUOTA,
    normalize, is_fragment, is_korean,
    load_ocr, extract_certified, has_disclaimer,
)

ERROR, WARN = "ERROR", "WARN"

# 시트 문장 대비 한 상품이 이보다 많은 행을 차지하면 다양성 경고.
PER_PRODUCT_SOFT_CAP = 8
# 문장으로 인정할 최소 정규화 길이(짧은 조각 컷).
MIN_SENTENCE_LEN = 8


class Report:
    """검사 결과 수집기. (심각도, 코드, 메시지)를 쌓는다."""

    def __init__(self):
        self.items: list[tuple[str, str, str]] = []

    def add(self, severity: str, code: str, msg: str) -> None:
        self.items.append((severity, code, msg))

    @property
    def errors(self) -> list:
        return [i for i in self.items if i[0] == ERROR]

    @property
    def warns(self) -> list:
        return [i for i in self.items if i[0] == WARN]

    def print(self) -> None:
        if not self.items:
            print("✅ 검수 통과 — 결함 없음")
            return
        for sev, code, msg in self.items:
            mark = "❌" if sev == ERROR else "⚠️ "
            print(f"{mark} [{code}] {msg}")
        print(f"\n=== ERROR {len(self.errors)} · WARN {len(self.warns)} ===")


# ── 순수 검사 함수들 (유닛테스트 대상) ──────────────────────────────

def check_required(rows: list[dict], tag: str, rep: Report) -> None:
    """필수 칸(product_id·sentence·product_type)이 비면 ERROR."""
    for r in rows:
        for col in ("product_id", "sentence", "product_type"):
            if not str(r.get(col, "")).strip():
                rep.add(ERROR, "required",
                        f"{tag} {r.get('id','?')}: {col} 비어 있음")


def check_label_blank(rows: list[dict], tag: str, rep: Report) -> None:
    """배포본의 label 칸은 반드시 비어 있어야 한다(내 힌트가 새면 정답 유출)."""
    leaked = [r.get("id", "?") for r in rows if str(r.get("label", "")).strip()]
    if leaked:
        rep.add(ERROR, "label-leak",
                f"{tag}: label 칸이 채워진 행 {len(leaked)}개 — 정답 유출 "
                f"({', '.join(leaked[:5])})")


def check_sentence_quality(rows: list[dict], tag: str, rep: Report) -> None:
    """문장 조각·영문·너무 짧은 조각이 섞였는지."""
    for r in rows:
        s = str(r.get("sentence", "")).strip()
        rid = r.get("id", "?")
        if len(normalize(s)) < MIN_SENTENCE_LEN:
            rep.add(ERROR, "too-short", f"{tag} {rid}: 너무 짧음 «{s}»")
        elif not is_korean(s):
            rep.add(ERROR, "not-korean", f"{tag} {rid}: 한국어 아님 «{s[:40]}»")
        elif is_fragment(s):
            rep.add(ERROR, "fragment", f"{tag} {rid}: 문장 조각 «{s[:40]}»")


def check_label_values(rows: list[dict], tag: str, rep: Report) -> None:
    """채워진 label이 있다면(내부 검수 파일) 허용 라벨인지."""
    allowed = set(LABELS)
    for r in rows:
        v = str(r.get("label", "")).strip()
        if v and v not in allowed:
            rep.add(ERROR, "bad-label",
                    f"{tag} {r.get('id','?')}: 허용 안 된 라벨 «{v}»")


def check_id_scheme(rows: list[dict], tag: str, rep: Report) -> None:
    """id 접두사와 is_cross_check가 일관적인지.

    S### = 공통(교차검증) → is_cross_check Y / 그 외 → N.
    """
    for r in rows:
        rid = str(r.get("id", ""))
        cc = str(r.get("is_cross_check", "")).strip()
        shared = rid.startswith("S")
        if shared and cc != "Y":
            rep.add(ERROR, "cross-flag",
                    f"{tag} {rid}: 공통행인데 is_cross_check≠Y ({cc})")
        if not shared and cc == "Y":
            rep.add(ERROR, "cross-flag",
                    f"{tag} {rid}: 고유행인데 is_cross_check=Y")


def check_product_type(rows: list[dict], tag: str, rep: Report) -> None:
    """product_type 불명(WARN)과 인정문구-유형 모순(WARN)을 본다."""
    unknown = [r.get("id", "?") for r in rows
               if str(r.get("product_type", "")).strip() == "불명"]
    if unknown:
        rep.add(WARN, "type-unknown",
                f"{tag}: product_type 불명 {len(unknown)}행 — 라벨러 확인 필요 "
                f"({', '.join(unknown[:5])})")
    for r in rows:
        cert = str(r.get("certified_function", "")).strip()
        ptype = str(r.get("product_type", "")).strip()
        # 인정 기능성 문구가 있는데 일반식품이면 판별 충돌 가능성.
        if cert and ptype == "일반식품":
            rep.add(WARN, "type-conflict",
                    f"{tag} {r.get('id','?')}: 인정문구 있는데 일반식품")


def check_per_product(rows: list[dict], tag: str, rep: Report) -> None:
    """한 상품이 시트를 과점하는지(다양성)."""
    c = Counter(r.get("product_id") for r in rows)
    for pid, n in c.most_common(3):
        if n > PER_PRODUCT_SOFT_CAP:
            rep.add(WARN, "product-skew",
                    f"{tag}: 상품 {pid} 이 {n}행 점유 (>{PER_PRODUCT_SOFT_CAP})")


def check_duplicates(sheets: dict[str, list[dict]], rep: Report) -> None:
    """비공통 문장이 중복되면 ERROR. 공통(S###)은 두 시트에 같아야 정상."""
    seen: dict[str, str] = {}
    for tag, rows in sheets.items():
        for r in rows:
            if str(r.get("id", "")).startswith("S"):
                continue  # 공통행은 의도적 중복
            key = normalize(str(r.get("sentence", "")))
            if key in seen:
                rep.add(ERROR, "dup-sentence",
                        f"중복 문장: {seen[key]} ↔ {tag} {r.get('id','?')} "
                        f"«{r.get('sentence','')[:36]}»")
            else:
                seen[key] = f"{tag} {r.get('id','?')}"


def check_shared_block(sheets: dict[str, list[dict]], rep: Report) -> None:
    """공통 S### 행이 모든 시트에서 같은 문장인지."""
    if len(sheets) < 2:
        return
    by_id: dict[str, dict[str, str]] = defaultdict(dict)
    for tag, rows in sheets.items():
        for r in rows:
            rid = str(r.get("id", ""))
            if rid.startswith("S"):
                by_id[rid][tag] = normalize(str(r.get("sentence", "")))
    tags = list(sheets)
    for rid, per_tag in by_id.items():
        vals = {per_tag.get(t) for t in tags}
        if len(vals) > 1 or None in vals:
            rep.add(ERROR, "shared-mismatch",
                    f"공통행 {rid}: 시트마다 문장이 다름")


def check_quota(sheets: dict[str, list[dict]], rep: Report) -> None:
    """힌트 기준 유형별 확보량이 목표에 못 미치면 WARN.

    공통행은 한 번만 센다(두 시트에 중복 등장하므로).
    """
    counted, seen_shared = Counter(), set()
    for rows in sheets.values():
        for r in rows:
            rid = str(r.get("id", ""))
            if rid.startswith("S"):
                if rid in seen_shared:
                    continue
                seen_shared.add(rid)
            h = r.get("hint")
            if h:
                counted[h] += 1
    for label, target in QUOTA.items():
        got = counted.get(label, 0)
        if got < target:
            rep.add(WARN, "quota",
                    f"쿼터 미달 {label}: {got}/{target}")


# ── 조립 ────────────────────────────────────────────────────────

def validate_rows(sheets: dict[str, list[dict]]) -> Report:
    """빌드 중 in-memory 행(힌트 포함)을 검사한다."""
    rep = Report()
    for tag, rows in sheets.items():
        check_required(rows, tag, rep)
        check_label_blank(rows, tag, rep)
        check_label_values(rows, tag, rep)
        check_sentence_quality(rows, tag, rep)
        check_id_scheme(rows, tag, rep)
        check_product_type(rows, tag, rep)
        check_per_product(rows, tag, rep)
    check_duplicates(sheets, rep)
    check_shared_block(sheets, rep)
    check_quota(sheets, rep)
    return rep


def read_xlsx(path: Path) -> list[dict]:
    """배포 xlsx의 라벨링 시트를 행 dict 리스트로 읽는다."""
    from openpyxl import load_workbook

    ws = load_workbook(path, data_only=True)["라벨링"]
    header = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append({h: ("" if v is None else v) for h, v in zip(header, r)})
    return rows


def check_xlsx_integrity(path: Path, rep: Report) -> None:
    """배포본 자체의 무결성: 헤더·드롭다운 존재."""
    from openpyxl import load_workbook

    wb = load_workbook(path)
    if "라벨링" not in wb.sheetnames:
        rep.add(ERROR, "no-sheet", f"{path.name}: '라벨링' 시트 없음")
        return
    ws = wb["라벨링"]
    header = [c.value for c in ws[1]]
    if header != COLUMNS:
        rep.add(ERROR, "header",
                f"{path.name}: 헤더 불일치\n    기대: {COLUMNS}\n    실제: {header}")
    dv_cols = {rng.split("$")[1] if "$" in rng else rng[0]
               for dv in ws.data_validations.dataValidation
               for rng in str(dv.sqref).split()}
    for col in ("label", "confidence", "product_type"):
        letter = chr(ord("A") + COLUMNS.index(col))
        if letter not in dv_cols:
            rep.add(ERROR, "no-dropdown",
                    f"{path.name}: {col}({letter}열) 드롭다운 없음")


def load_master_hints(master: Path) -> dict[str, dict[str, str]]:
    """master jsonl에서 (sheet,id)→hint 를 읽어 xlsx 행에 붙인다."""
    import json
    out = defaultdict(dict)
    if not master.exists():
        return out
    for line in open(master):
        if line.strip():
            r = json.loads(line)
            out[r.get("sheet", "?")][r.get("id")] = r.get("hint")
    return out


def main() -> int:
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--a", default="data/holdout_A.xlsx")
    ap.add_argument("--b", default="data/holdout_B.xlsx")
    ap.add_argument("--master", default="data/holdout_master_v1.jsonl")
    args = ap.parse_args()

    hints = load_master_hints(Path(args.master))
    sheets = {}
    rep = Report()
    for tag, path in (("A", Path(args.a)), ("B", Path(args.b))):
        if not path.exists():
            rep.add(ERROR, "no-file", f"{path} 없음")
            continue
        rows = read_xlsx(path)
        for r in rows:  # 힌트를 붙여 쿼터 검사에 쓴다
            r["hint"] = hints.get(tag, {}).get(str(r.get("id", "")))
        sheets[tag] = rows
        check_xlsx_integrity(path, rep)

    if sheets:
        rows_rep = validate_rows(sheets)
        rep.items = rows_rep.items + rep.items

    print(f"=== 홀드아웃 검수: A {len(sheets.get('A', []))}행 · "
          f"B {len(sheets.get('B', []))}행 ===")
    rep.print()
    return 1 if rep.errors else 0


if __name__ == "__main__":
    sys.exit(main())
