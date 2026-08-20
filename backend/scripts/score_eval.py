# -*- coding: utf-8 -*-
"""평가셋 채점: 사람 정답 라벨 vs LLM 판정 비교.

대수가 라벨링한 `data/cosmetic_eval_labeling.xlsx`(라벨 칸 D열)을 읽어,
같은 문장을 LLM(Gemini)에 판정시키고 정답과 대조해 점수를 낸다.

실행(backend/에서):
  ./venv/bin/python scripts/score_eval.py            # 실채점(Gemini 호출)
  ./venv/bin/python scripts/score_eval.py --dry       # 호출 없이 배선만 점검

미탐(위반을 합법/대상외로 놓침)이 1급 지표다.
"""
import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, "src")
from barum.judge.cosmetic import JUDGE_PROMPT  # noqa: E402
from barum.vlm import get_vlm  # noqa: E402

XLSX = Path("data/cosmetic_eval_labeling.xlsx")       # ver1(유형 1축, 낡음. 보존용)
XLSX_V2 = Path("data/cosmetic_eval_labeling_v2.xlsx")  # ver2(유형+확정도 2축, 현행)
COMPARE = Path("data/eval_compare.csv")   # 모델별 요약 누적 → 비교표

LABELS = ["합법", "1호_의약품오인", "2호_기능성오인", "5호_거짓과장기만", "대상외"]
VIOLATION = {"1호_의약품오인", "2호_기능성오인", "5호_거짓과장기만"}

# 판정 프롬프트는 barum.judge.cosmetic이 원본. 판정기(PromptJudge)와 채점기가 같은
# 프롬프트를 써야 채점 결과가 실제 API 판정과 일치한다.


def load_labeled():
    """평가셋에서 (번호, 문장, 사람라벨, 확정도)를 읽는다.

    **ver2가 있으면 ver2를 쓴다.** ver1은 위반유형 1축뿐이라 "검토필요"를 적을 칸이
    없어, 검토필요여야 할 문장이 전부 합법 아니면 위반유형으로 몰려 있었다. 실측으로
    ver1과 현행 963셋에 겹치는 35문장 중 14건(40%)이 합법 여부부터 어긋났다
    (2026-08-20, `docs/result/2026-08-20_판정로직_고도화_로그.md` ⑥). ver2는
    `build_eval_goldset_v2.py`가 현행 정답셋 기준으로 만든다.

    `flag`는 ver2에만 있다(위반·검토필요, 합법/대상외/애매는 ""). ver1로 떨어지면
    빈 문자열이라 기존 1축 채점 그대로 동작한다.

    ver2에서 검토필요 행은 `human`(유형)이 비어 있을 수 있다 — 963셋이 검토필요
    160건 중 154건에 유형을 안 매겼기 때문이다("유형은 정해도 확정을 못 하는" 상태).
    그런 행은 유형 불문 플래그 여부만 보는 게 맞다(채점기가 그렇게 처리한다).
    """
    path = XLSX_V2 if XLSX_V2.exists() else XLSX
    if not path.exists():
        sys.exit(f"[없음] {path} 먼저 준비할 것 (ver2는 build_eval_goldset_v2.py로 생성)")
    if path == XLSX:
        print(f"[주의] ver2({XLSX_V2})가 없어 ver1로 채점한다. ver1은 검토필요 축이 "
              f"없어 지표를 신뢰할 수 없다 — build_eval_goldset_v2.py를 먼저 돌릴 것.")
    wb = openpyxl.load_workbook(path)
    ws = wb["라벨링"]
    headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    flag_col = headers.get("확정도")  # ver2에만 있다. 고정 열번호로 읽지 않는다.
    rows = []
    for r in range(2, ws.max_row + 1):
        n = ws.cell(r, 1).value
        text = ws.cell(r, 3).value
        human = (ws.cell(r, 4).value or "").strip()
        flag = (ws.cell(r, flag_col).value or "").strip() if flag_col else ""
        if not text:
            continue
        src = str(ws.cell(r, 2).value or "").strip()  # "24500688/detail_003"
        rows.append({"n": n, "text": text, "human": human, "flag": flag,
                     "product": src.split("/")[0] if "/" in src else ""})
    return rows


def judge_batch(vlm, batch):
    """문장 배치를 LLM에 판정시켜 번호→(라벨,근거) dict 반환.

    모델이 가끔 {"results":[...]} 대신 통짜 리스트를 뱉는다. 그때 res.get()이
    AttributeError를 던져 채점 전체가 죽던 걸(핸드오프 §2.3) 예상된 실패로 흡수한다.
    이 배치는 빈 결과로 두면 되고(호출자가 "(없음)"으로 처리), 다음 배치는 계속 돈다.
    PromptJudge는 이미 같은 방식으로 고쳐져 있다.
    """
    items = "\n".join(f'{b["n"]}. {b["text"]}' for b in batch)
    res = vlm.generate_json(JUDGE_PROMPT.format(items=items), [])
    if not isinstance(res, dict):
        print(f"    [skip] judge 배치 {batch[0]['n']}~{batch[-1]['n']}: "
              f"dict 아님({type(res).__name__}) — 이 배치 미판정")
        return {}
    out = {}
    for item in res.get("results", []):
        try:
            out[int(item["n"])] = (item.get("label", "").strip(), item.get("reason", ""))
        except (KeyError, ValueError, TypeError):
            continue
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--provider", default="gemini", help="gemini | openai")
    ap.add_argument("--model", default=None, help="모델명(생략 시 provider 기본값)")
    ap.add_argument("--dry", action="store_true", help="API 호출 없이 배선만 점검")
    ap.add_argument("--batch", type=int, default=12)
    args = ap.parse_args()

    rows = load_labeled()
    scored = [r for r in rows if r["human"] in LABELS]      # 채점 대상(유효 라벨)
    pending = [r for r in rows if r["human"] == ""]          # 라벨 미완
    abstain = [r for r in rows if r["human"] and r["human"] not in LABELS]  # 애매 등

    print(f"전체 {len(rows)}문장 / 라벨 완료 {len(rows)-len(pending)} / 미완 {len(pending)} / "
          f"채점대상(유효라벨) {len(scored)} / 제외(애매 등) {len(abstain)}")

    if args.dry:
        print("\n[--dry] Gemini 호출 안 함. 첫 배치 판정 프롬프트 미리보기:\n")
        preview = scored[:args.batch] or rows[:args.batch]
        items = "\n".join(f'{b["n"]}. {b["text"]}' for b in preview)
        print(JUDGE_PROMPT.format(items=items)[:1200], "...")
        return

    if not scored:
        sys.exit("\n채점할 라벨이 없다. 대수 라벨링(D열) 후 다시 실행할 것.")

    try:
        vlm = get_vlm(args.provider, model=args.model)
    except RuntimeError as e:
        sys.exit(f"[키 없음] {e}")
    print(f"provider={args.provider}  모델={vlm.model}\n", flush=True)
    ai = {}
    for i in range(0, len(scored), args.batch):
        batch = scored[i:i + args.batch]
        ai.update(judge_batch(vlm, batch))
        print(f"  판정 {min(i+args.batch, len(scored))}/{len(scored)}", flush=True)

    # 채점
    match = miss = false_alarm = 0
    misses, false_alarms, unknown = [], [], []
    wb = openpyxl.Workbook(); r_ws = wb.active; r_ws.title = "결과"
    r_ws.append(["번호", "문장", "사람", "AI", "AI근거", "일치"])
    for r in scored:
        human = r["human"]
        ai_label, reason = ai.get(r["n"], ("(없음)", ""))
        ok = (ai_label == human)
        match += ok
        if ai_label not in LABELS:
            unknown.append(r["n"])
        if human in VIOLATION and ai_label in {"합법", "대상외"}:
            miss += 1; misses.append((r["n"], r["text"], human, ai_label))
        if human == "합법" and ai_label in VIOLATION:
            false_alarm += 1; false_alarms.append((r["n"], r["text"], ai_label))
        r_ws.append([r["n"], r["text"], human, ai_label, reason, "O" if ok else "X"])
    safe_model = (vlm.model or "default").replace("/", "-")
    out = Path(f"data/eval_result_{args.provider}_{safe_model}.xlsx")
    wb.save(out)

    n = len(scored)
    acc = match / n * 100
    print("\n" + "=" * 46)
    print(f"provider={args.provider}  모델={vlm.model}")
    print(f"채점 대상: {n}문장")
    print(f"전체 일치율: {match}/{n} = {acc:.1f}%")
    print(f"미탐(위반→합법/대상외, 1급): {miss}건  ← 낮을수록 좋음")
    print(f"오탐(합법→위반): {false_alarm}건")
    if unknown:
        print(f"AI가 규격 밖 라벨 뱉음: {len(unknown)}건 {unknown}")
    if misses:
        print("\n[미탐 목록 — 제일 중요]")
        for n_, t, h, a in misses:
            print(f"  #{n_} 사람={h} AI={a} | {t[:40]}")
    if false_alarms:
        print("\n[오탐 목록]")
        for n_, t, a in false_alarms:
            print(f"  #{n_} AI={a} | {t[:40]}")

    # 모델별 요약을 한 파일에 누적 → 비교표(Gemini vs OpenAI 나란히)
    new_file = not COMPARE.exists()
    with open(COMPARE, "a", newline="") as f:
        w = csv.writer(f)
        if new_file:
            w.writerow(["시각", "provider", "모델", "채점수", "일치율%", "미탐", "오탐", "토큰"])
        w.writerow([datetime.now().strftime("%m-%d %H:%M"), args.provider, vlm.model,
                    n, f"{acc:.1f}", miss, false_alarm, vlm.total_tokens])

    print("=" * 46)
    print(f"상세: {out}")
    print(f"비교표(누적): {COMPARE}")
    print(f"누적 토큰: {vlm.total_tokens:,}")


if __name__ == "__main__":
    main()
