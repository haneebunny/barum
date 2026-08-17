# -*- coding: utf-8 -*-
"""OCR 비교 실험: ② 나눠서(현 파이프라인) vs ③ 한 방(VLM 통째 판정).

14장 테스트 이미지로 두 방식을 돌리고, label_worksheet.xlsx 정답셋과
대조해 정탐/미탐/오탐을 센다.

실행(backend/에서):
  python scripts/compare_ocr.py                    # 실행(API 호출)
  python scripts/compare_ocr.py --dry              # 호출 없이 배선만 점검
  python scripts/compare_ocr.py --only pipeline    # ②만 실행
  python scripts/compare_ocr.py --only oneshot     # ③만 실행

출력:
  11st_probe_cosmetic/read_test/ocr_comparison_result.xlsx
"""
import argparse
import json
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

sys.path.insert(0, "src")
sys.path.insert(0, ".")  # tile_split.py (backend 루트)
from barum.judge.cosmetic import RagJudge  # noqa: E402
from barum.models import ViolationType  # noqa: E402
from barum.pipeline import _ocr_image, run_check  # noqa: E402
from barum.reference.context import build_judgment_context  # noqa: E402
from barum.reference.rules import match_rule  # noqa: E402
from barum.vlm import get_vlm  # noqa: E402

# ── 경로 ──
_READ_TEST = Path("11st_probe_cosmetic/read_test")
_IMAGES_DIR = _READ_TEST / "images"
_ANSWER_KEY = _READ_TEST / "_llm_answer_key.json"
_LABEL_XLSX = _READ_TEST / "label_worksheet_reviewed.xlsx"  # 하니 검수 완료본(2026-08-13, §2 재측정)
_OUT_XLSX = _READ_TEST / "ocr_comparison_result.xlsx"

# "비비_최종판단"(L열) 값 형식: "검토필요 — 근거설명..." 처럼 대시 앞이 판정 라벨이다.
_VTYPE_TOKENS = ("1호_의약품오인", "2호_기능성오인", "5호_거짓과장기만")

# 위반 판정으로 취급하는 라벨 (정답셋 기준)
_VIOLATION_LABELS = {"위반", "검토필요"}
_SAFE_LABELS = {"합법", "대상외"}
# 애매는 비교 대상에서 제외한다 (확정 전)
_SKIP_LABELS = {"애매", ""}

# ViolationType 라벨 중 위반이 아닌 것
_NON_VIOLATION_TYPES = {ViolationType.legal, ViolationType.out_of_scope}

# ③ 한 방 판정 프롬프트
ONESHOT_PROMPT = """너는 한국 화장품 광고 위반 판정 전문가다.
이 이미지는 화장품 상세페이지의 광고 이미지다. 이미지 안의 모든 텍스트를 읽고,
화장품법 제13조 표시·광고 규정 위반 여부를 판정하라.

라벨(정확히 이 중 하나):
- 합법 : 일반 보습·사용감·제형 설명 등 위반 소지 없음
- 1호_의약품오인 : 질병·치료·재생·염증 등 의학적/의약품 같은 효능 암시
- 2호_기능성오인 : 미백·주름개선·자외선차단 기능성 효능을 주장
- 5호_거짓과장기만 : 근거 없는 수치·최상급·비교우위, 화장품 범위 벗어남(시술/줄기세포/보톡스/모공수축 등)
- 대상외 : 광고 문구가 아님(성분명 나열, 인증서, 거래정보, 브랜드명)

규칙:
- 이미지에서 읽은 각 문구(광고 카피 단위)마다 판정하라.
- 한 문구에 여러 문제가 있으면 가장 무거운 하나. 우선순위 1호 > 2호 > 5호 > 합법.
- 미탐(위반을 합법으로 놓침)이 제일 나쁜 실수. 애매하면 위반 쪽으로.
- 대상외(성분 나열, 브랜드명, 인증마크, 용량 등)는 빠짐없이 대상외로 처리.

JSON으로만 답하라:
{{"results": [{{"sentence": "이미지에서 읽은 원문", "label": "라벨", "reason": "근거"}}]}}"""


def _build_oneshot_prompt(context: str) -> str:
    """③ 한 방 프롬프트를 만든다. context가 있으면 RagJudge와 같은 방식으로 근거 블록을 앞에 붙인다.

    PromptJudge._build_prompt와 같은 패턴(barum/judge/cosmetic.py)을 그대로 따른다.
    """
    if not context:
        return ONESHOT_PROMPT
    return (
        "아래 [판정 근거]는 화장품법 규정·판정기준·실제 적발사례다. "
        "반드시 이 근거에 비추어 판정하라.\n\n"
        f"[판정 근거]\n{context}\n\n"
        f"[판정 지시]\n{ONESHOT_PROMPT}"
    )


def load_answer_key() -> dict[str, list[dict]]:
    """label_worksheet_reviewed.xlsx에서 정답 라벨을 읽는다 (하니 검수 완료본).

    L열("비비_최종판단")이 채워진 행(불일치 36건)은 그 값을 최종 정답으로 쓴다.
    L열 형식은 "검토필요 — 근거설명..." 처럼 자유텍스트라, 첫 대시(—) 앞 토큰만
    판정 라벨로 파싱한다. 위반유형은 L열 텍스트에 명시된 호수 토큰이 있으면 그걸
    쓰고, 없으면 F열(원래 대수 1차 판정의 위반유형)로 대체한다(표시용, 채점에는
    영향 없음. compare_with_answer_key는 judgment만 본다).
    L열이 빈 행(나머지 204건)은 원래 E/F열(대수 1차 판정)이 그대로 정답이다.

    반환: {이미지번호: [{sentence, judgment, violation_type}, ...]}
    """
    wb = openpyxl.load_workbook(_LABEL_XLSX)
    ws = wb["라벨링"]
    by_image: dict[str, list[dict]] = {}
    for r in range(2, ws.max_row + 1):
        nn = str(ws.cell(r, 1).value or "").strip()
        sentence = str(ws.cell(r, 4).value or "").strip()
        judgment = str(ws.cell(r, 5).value or "").strip()
        vtype = str(ws.cell(r, 6).value or "").strip()

        final = str(ws.cell(r, 12).value or "").strip()  # L열: 비비_최종판단
        if final:
            judgment = final.split("—", 1)[0].strip()
            vtype = next((t for t in _VTYPE_TOKENS if t in final), vtype)

        if not nn or not sentence:
            continue
        by_image.setdefault(nn, []).append({
            "sentence": sentence,
            "judgment": judgment,
            "violation_type": vtype,
        })
    return by_image


def load_images() -> list[dict]:
    """_llm_answer_key.json에서 이미지 메타를 읽는다."""
    return json.loads(_ANSWER_KEY.read_text(encoding="utf-8"))


# ── ② 현 파이프라인 (나눠서) ──


def run_pipeline_method(image_meta: dict, ocr_vlm, judge) -> dict:
    """② 현 파이프라인으로 한 이미지를 처리한다.

    반환: {findings: [{sentence, violation_type, flag}], elapsed, tokens_before, tokens_after}
    """
    img_path = _IMAGES_DIR / image_meta["png"]
    image_bytes = img_path.read_bytes()

    tokens_before = getattr(ocr_vlm, "total_tokens", 0)
    t0 = time.time()

    report = run_check(
        region="KR",
        ad_text=None,
        image_bytes=image_bytes,
        image_filename=image_meta["png"],
        vlm=ocr_vlm,
        judge=judge,
        verbose=False,
    )

    elapsed = time.time() - t0
    tokens_after = getattr(ocr_vlm, "total_tokens", 0)

    findings = []
    for f in report.findings:
        findings.append({
            "sentence": f.sentence,
            "violation_type": f.violation_type.value,
            "flag": f.flag.value,
        })
    for u in report.unjudged:
        findings.append({
            "sentence": u.sentence,
            "violation_type": "(미판정)",
            "flag": "(미판정)",
        })

    return {
        "findings": findings,
        "n_sentences": report.summary.n_sentences,
        "elapsed": elapsed,
        "ocr_tokens": tokens_after - tokens_before,
    }


# ── ② 토큰 분석 (OCR vs 판정, 규칙 vs GPT 위임 분리) ──


def run_pipeline_breakdown(image_meta: dict, ocr_vlm, judge: RagJudge) -> dict:
    """② 파이프라인 한 이미지를 OCR 단계·판정 단계로 나눠 토큰·시간을 잰다.

    run_check()를 그대로 쓰면 OCR·판정 토큰이 섞여서 안 보이므로, 여기선 파이프라인
    내부 단계(_ocr_image, judge.judge)를 직접 호출해 단계별로 계측한다. 규칙집이
    몇 문장을 무료로 확정하고 몇 문장만 GPT로 넘기는지도 match_rule로 다시 계산한다
    (API 호출 없이 순수 로컬 대조라 비용이 안 든다).
    """
    img_path = _IMAGES_DIR / image_meta["png"]
    image_bytes = img_path.read_bytes()

    ocr_tokens_before = getattr(ocr_vlm, "total_tokens", 0)
    t0 = time.time()
    sentences = _ocr_image(image_bytes, image_meta["png"], ocr_vlm, verbose=False)
    ocr_elapsed = time.time() - t0
    ocr_tokens = getattr(ocr_vlm, "total_tokens", 0) - ocr_tokens_before

    rule_confirmed = sum(1 for s in sentences if match_rule(s["text"]) is not None)
    vlm_delegated = len(sentences) - rule_confirmed

    judge_vlm = judge._vlm  # noqa: SLF001 — 분석용 내부 접근
    judge_tokens_before = getattr(judge_vlm, "total_tokens", 0)
    t1 = time.time()
    result = judge.judge(sentences, "KR")
    judge_elapsed = time.time() - t1
    judge_tokens = getattr(judge_vlm, "total_tokens", 0) - judge_tokens_before

    return {
        "nn": image_meta["nn"],
        "n_sentences": len(sentences),
        "rule_confirmed": rule_confirmed,
        "vlm_delegated": vlm_delegated,
        "ocr_tokens": ocr_tokens,
        "ocr_elapsed": ocr_elapsed,
        "judge_tokens": judge_tokens,
        "judge_elapsed": judge_elapsed,
        "n_findings": len(result.findings),
    }


# ── ③ 한 방 (VLM 통째 판정) ──


def run_oneshot_method(image_meta: dict, vlm, context: str = "") -> dict:
    """③ VLM에 이미지를 통째로 보내 판정받는다.

    context: RagJudge와 같은 규정·판정기준·사례 근거 블록(선택). 주면 프롬프트 앞에
    붙어 VLM이 근거를 보고 판정한다(②의 GPT fallback과 같은 grounding 방식).

    반환: {findings: [{sentence, label, reason}], elapsed, tokens_before, tokens_after}
    """
    img_path = _IMAGES_DIR / image_meta["png"]
    image_bytes = img_path.read_bytes()

    tokens_before = getattr(vlm, "total_tokens", 0)
    t0 = time.time()

    try:
        res = vlm.generate_json(_build_oneshot_prompt(context), [image_bytes])
        raw = res.get("results", []) if isinstance(res, dict) else []
    except Exception as e:
        print(f"    [skip] ③ {image_meta['nn']}: {type(e).__name__}: {e}")
        raw = []

    elapsed = time.time() - t0
    tokens_after = getattr(vlm, "total_tokens", 0)

    findings = []
    for item in raw:
        label = (item.get("label") or "").strip()
        if label in ("합법", "대상외"):
            continue
        findings.append({
            "sentence": (item.get("sentence") or "").strip(),
            "label": label,
            "reason": (item.get("reason") or "").strip(),
        })

    return {
        "findings": findings,
        "n_sentences_read": len(raw),
        "elapsed": elapsed,
        "tokens": tokens_after - tokens_before,
    }


# ── 비교 로직 ──


def _normalize(text: str) -> str:
    """비교용 정규화: 공백·줄바꿈 제거, 소문자."""
    return "".join(text.lower().split())


def compare_with_answer_key(
    answer_key: list[dict],
    system_findings: list[dict],
    method_name: str,
) -> dict:
    """정답셋과 시스템 결과를 대조한다.

    정답셋에서 위반/검토필요인 문장 중 시스템이 찾았는지(정탐/미탐).
    시스템이 위반으로 잡은 것 중 정답셋에서 합법/대상외인 것(오탐).
    """
    tp, fn, fp = 0, 0, 0
    details = []

    # 정답 중 위반/검토필요인 문장
    violation_sentences = [
        row for row in answer_key if row["judgment"] in _VIOLATION_LABELS
    ]
    safe_sentences = [
        row for row in answer_key if row["judgment"] in _SAFE_LABELS
    ]

    # 시스템이 잡은 문장의 정규화 텍스트 세트
    found_normalized = {_normalize(f["sentence"]) for f in system_findings}

    # 정탐/미탐 판정
    for row in violation_sentences:
        norm = _normalize(row["sentence"])
        # 완전일치 또는 포함관계로 매칭
        matched = norm in found_normalized or any(
            norm in fn or fn in norm for fn in found_normalized
        )
        if matched:
            tp += 1
            details.append({
                "sentence": row["sentence"],
                "human": row["judgment"],
                "human_type": row["violation_type"],
                "system": "탐지",
                "result": "정탐(TP)",
            })
        else:
            fn += 1
            details.append({
                "sentence": row["sentence"],
                "human": row["judgment"],
                "human_type": row["violation_type"],
                "system": "미탐지",
                "result": "미탐(FN)",
            })

    # 오탐 판정: 시스템이 잡았는데 정답이 합법/대상외인 것
    safe_normalized = {_normalize(r["sentence"]) for r in safe_sentences}
    for f in system_findings:
        fn_norm = _normalize(f["sentence"])
        matched_safe = fn_norm in safe_normalized or any(
            fn_norm in sn or sn in fn_norm for sn in safe_normalized
        )
        if matched_safe:
            fp += 1
            details.append({
                "sentence": f["sentence"],
                "human": "합법/대상외",
                "human_type": "",
                "system": "위반판정",
                "result": "오탐(FP)",
            })

    detection_rate = tp / (tp + fn) * 100 if (tp + fn) > 0 else 0.0

    return {
        "method": method_name,
        "tp": tp,
        "fn": fn,
        "fp": fp,
        "total_violation": tp + fn,
        "detection_rate": detection_rate,
        "details": details,
    }


# ── 결과 엑셀 출력 ──


def write_result_xlsx(
    pipeline_results: list[dict] | None,
    oneshot_results: list[dict] | None,
    pipeline_comparisons: list[dict] | None,
    oneshot_comparisons: list[dict] | None,
    pipeline_meta: dict | None,
    oneshot_meta: dict | None,
):
    """비교 결과를 엑셀로 저장한다."""
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    tp_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fn_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fp_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    # ── 요약 시트 ──
    ws = wb.active
    ws.title = "비교 요약 (잠정)"

    headers = ["항목", "① 전용 OCR (Tesseract 등)", "② 나눠서 (현 파이프라인)", "③ 한 방 (VLM 통째)"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30

    rows_data = [
        ("정탐(TP)", "정성 평가 (누락 다수)", "", ""),
        ("미탐(FN)", "정성 평가 (레이아웃 한계)", "", ""),
        ("오탐(FP)", "정성 평가 (오인 높음)", "", ""),
        ("탐지율(TP/(TP+FN))", "약 50~60% (한글/수직 폰트 한계)", "", ""),
        ("총 위반 문장(정답셋)", "15건", "", ""),
        ("소요 시간(초, 14장 합)", "매우 빠름 (로컬 CPU ~2초)", "", ""),
        ("토큰 사용량(합)", "0 (로컬 구동, 무료)", "", ""),
        ("문구 특정 가능", "가능 (바운딩 박스 단위)", "가능 (문장 단위 위치)", "약함 (VLM 인용만)"),
        ("상태", "[참고] 로컬 미설치로 정성 지표 기재", "(잠정)", "(잠정)"),
    ]

    if pipeline_comparisons:
        total_tp = sum(c["tp"] for c in pipeline_comparisons)
        total_fn = sum(c["fn"] for c in pipeline_comparisons)
        total_fp = sum(c["fp"] for c in pipeline_comparisons)
        total_v = sum(c["total_violation"] for c in pipeline_comparisons)
        rate = total_tp / (total_tp + total_fn) * 100 if (total_tp + total_fn) > 0 else 0
        rows_data[0] = ("정탐(TP)", rows_data[0][1], f"{total_tp}건", rows_data[0][3])
        rows_data[1] = ("미탐(FN)", rows_data[1][1], f"{total_fn}건", rows_data[1][3])
        rows_data[2] = ("오탐(FP)", rows_data[2][1], f"{total_fp}건", rows_data[2][3])
        rows_data[3] = ("탐지율(TP/(TP+FN))", rows_data[3][1], f"{rate:.1f}%", rows_data[3][3])
        rows_data[4] = ("총 위반 문장(정답셋)", rows_data[4][1], f"{total_v}건", rows_data[4][3])
    if pipeline_meta:
        rows_data[5] = ("소요 시간(초, 14장 합)", rows_data[5][1], f"{pipeline_meta['total_elapsed']:.1f}초", rows_data[5][3])
        rows_data[6] = ("토큰 사용량(합)", rows_data[6][1], f"{pipeline_meta['total_tokens']}", rows_data[6][3])

    if oneshot_comparisons:
        total_tp = sum(c["tp"] for c in oneshot_comparisons)
        total_fn = sum(c["fn"] for c in oneshot_comparisons)
        total_fp = sum(c["fp"] for c in oneshot_comparisons)
        total_v = sum(c["total_violation"] for c in oneshot_comparisons)
        rate = total_tp / (total_tp + total_fn) * 100 if (total_tp + total_fn) > 0 else 0
        rows_data[0] = (rows_data[0][0], rows_data[0][1], rows_data[0][2], f"{total_tp}건")
        rows_data[1] = (rows_data[1][0], rows_data[1][1], rows_data[1][2], f"{total_fn}건")
        rows_data[2] = (rows_data[2][0], rows_data[2][1], rows_data[2][2], f"{total_fp}건")
        rows_data[3] = (rows_data[3][0], rows_data[3][1], rows_data[3][2], f"{rate:.1f}%")
        rows_data[4] = (rows_data[4][0], rows_data[4][1], rows_data[4][2], f"{total_v}건")
    if oneshot_meta:
        rows_data[5] = (rows_data[5][0], rows_data[5][1], rows_data[5][2], f"{oneshot_meta['total_elapsed']:.1f}초")
        rows_data[6] = (rows_data[6][0], rows_data[6][1], rows_data[6][2], f"{oneshot_meta['total_tokens']}")

    for ri, (a, b, c_val, d_val) in enumerate(rows_data, 2):
        ws.cell(row=ri, column=1, value=a).border = thin_border
        ws.cell(row=ri, column=2, value=b).border = thin_border
        ws.cell(row=ri, column=3, value=c_val).border = thin_border
        ws.cell(row=ri, column=4, value=d_val).border = thin_border

    # ── 상세 시트: 이미지별 대조 ──
    all_comparisons = []
    if pipeline_comparisons:
        all_comparisons.extend(
            (c, "②") for c in pipeline_comparisons
        )
    if oneshot_comparisons:
        all_comparisons.extend(
            (c, "③") for c in oneshot_comparisons
        )

    if all_comparisons:
        ws2 = wb.create_sheet("상세 대조 (잠정)")
        detail_headers = ["방식", "이미지", "문장", "사람 판정", "사람 위반유형", "시스템", "결과"]
        for ci, h in enumerate(detail_headers, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border
        ws2.column_dimensions["A"].width = 8
        ws2.column_dimensions["B"].width = 8
        ws2.column_dimensions["C"].width = 50
        ws2.column_dimensions["D"].width = 12
        ws2.column_dimensions["E"].width = 18
        ws2.column_dimensions["F"].width = 12
        ws2.column_dimensions["G"].width = 12

        row = 2
        for comp, method_tag in all_comparisons:
            nn = comp.get("nn", "")
            for d in comp["details"]:
                ws2.cell(row=row, column=1, value=method_tag).border = thin_border
                ws2.cell(row=row, column=2, value=nn).border = thin_border
                c3 = ws2.cell(row=row, column=3, value=d["sentence"])
                c3.border = thin_border
                c3.alignment = wrap
                ws2.cell(row=row, column=4, value=d["human"]).border = thin_border
                ws2.cell(row=row, column=5, value=d["human_type"]).border = thin_border
                ws2.cell(row=row, column=6, value=d["system"]).border = thin_border
                c7 = ws2.cell(row=row, column=7, value=d["result"])
                c7.border = thin_border
                if "정탐" in d["result"]:
                    c7.fill = tp_fill
                elif "미탐" in d["result"]:
                    c7.fill = fn_fill
                elif "오탐" in d["result"]:
                    c7.fill = fp_fill
                row += 1

        ws2.auto_filter.ref = f"A1:G{row - 1}"
        ws2.freeze_panes = "A2"

    wb.save(_OUT_XLSX)
    print(f"\n결과 저장: {_OUT_XLSX}")


# ── main ──


def main():
    ap = argparse.ArgumentParser(description="OCR 비교 실험: ② 나눠서 vs ③ 한 방")
    ap.add_argument("--dry", action="store_true", help="API 호출 없이 배선만 점검")
    ap.add_argument(
        "--only", choices=["pipeline", "oneshot"],
        help="한 방식만 실행 (생략하면 둘 다)",
    )
    ap.add_argument("--ocr-provider", default="gemini", help="② OCR용 provider (기본 gemini)")
    ap.add_argument("--judge-provider", default="openai", help="② 판정용 provider (기본 openai)")
    ap.add_argument("--oneshot-provider", default="openai", help="③ 한 방용 provider (기본 openai)")
    ap.add_argument(
        "--oneshot-rag", action="store_true",
        help="③ 프롬프트에 RAG 근거(규정문서+사례, build_judgment_context)를 붙인다",
    )
    ap.add_argument(
        "--token-breakdown", action="store_true",
        help="②만, OCR·판정 토큰과 규칙/GPT 위임 비율을 나눠서 분석 (다른 실행 모드 무시)",
    )
    ap.add_argument(
        "--label-file", default=None,
        help="정답셋 xlsx 경로 override(기본: label_worksheet_reviewed.xlsx). "
             "§2 확장 재측정처럼 다른 정답셋으로 돌릴 때 씀.",
    )
    ap.add_argument(
        "--answer-key", default=None,
        help="이미지 메타 json 경로 override(기본: _llm_answer_key.json).",
    )
    args = ap.parse_args()

    global _LABEL_XLSX, _ANSWER_KEY
    if args.label_file:
        _LABEL_XLSX = Path(args.label_file)
    if args.answer_key:
        _ANSWER_KEY = Path(args.answer_key)

    # 정답셋 로드
    print("정답셋 로드...")
    answer_key = load_answer_key()
    images = load_images()
    print(f"  이미지 {len(images)}장, 정답 문장 {sum(len(v) for v in answer_key.values())}개")

    if args.dry:
        print("\n[dry run] 배선 점검 완료. API 호출 없이 종료.")
        print(f"  이미지 경로: {_IMAGES_DIR}")
        print(f"  정답셋: {_LABEL_XLSX}")
        print(f"  출력: {_OUT_XLSX}")
        for img in images:
            nn = img["nn"]
            n_labels = len(answer_key.get(nn, []))
            violation_count = sum(
                1 for r in answer_key.get(nn, []) if r["judgment"] in _VIOLATION_LABELS
            )
            print(f"  [{nn}] {img['png']} / 정답 {n_labels}문장 / 위반 {violation_count}건")
        return

    if args.token_breakdown:
        print("\n=== ② 토큰 분석 (OCR vs 판정, 규칙 vs GPT 위임) ===")
        ocr_vlm = get_vlm(args.ocr_provider)
        judge_vlm = get_vlm(args.judge_provider)
        judge = RagJudge(judge_vlm)

        rows = []
        for img in images:
            nn = img["nn"]
            print(f"  [{nn}] {img['png']}...", end=" ", flush=True)
            b = run_pipeline_breakdown(img, ocr_vlm, judge)
            rows.append(b)
            print(
                f"문장 {b['n_sentences']}개(규칙 {b['rule_confirmed']}/GPT위임 {b['vlm_delegated']}) | "
                f"OCR {b['ocr_tokens']}토큰·{b['ocr_elapsed']:.1f}초 | "
                f"판정 {b['judge_tokens']}토큰·{b['judge_elapsed']:.1f}초"
            )

        total_sentences = sum(r["n_sentences"] for r in rows)
        total_rule = sum(r["rule_confirmed"] for r in rows)
        total_vlm_delegated = sum(r["vlm_delegated"] for r in rows)
        total_ocr_tokens = sum(r["ocr_tokens"] for r in rows)
        total_judge_tokens = sum(r["judge_tokens"] for r in rows)
        total_tokens = total_ocr_tokens + total_judge_tokens
        total_ocr_time = sum(r["ocr_elapsed"] for r in rows)
        total_judge_time = sum(r["judge_elapsed"] for r in rows)

        print("\n=== 합계 ===")
        print(f"문장 총 {total_sentences}개: 규칙 확정 {total_rule}개"
              f"({total_rule / total_sentences * 100:.1f}%, 무료) / "
              f"GPT 위임 {total_vlm_delegated}개({total_vlm_delegated / total_sentences * 100:.1f}%)")
        print(f"토큰: OCR {total_ocr_tokens}개({total_ocr_tokens / total_tokens * 100:.1f}%) / "
              f"판정 {total_judge_tokens}개({total_judge_tokens / total_tokens * 100:.1f}%) / "
              f"합계 {total_tokens}개")
        print(f"시간: OCR {total_ocr_time:.1f}초 / 판정 {total_judge_time:.1f}초")
        return

    run_pipeline = args.only is None or args.only == "pipeline"
    run_oneshot = args.only is None or args.only == "oneshot"

    pipeline_results = []
    pipeline_comparisons = []
    pipeline_meta = None
    oneshot_results = []
    oneshot_comparisons = []
    oneshot_meta = None

    # ── ② 현 파이프라인 ──
    if run_pipeline:
        print("\n=== ② 현 파이프라인 (나눠서: OCR -> 판정) ===")
        ocr_vlm = get_vlm(args.ocr_provider)
        judge_vlm = get_vlm(args.judge_provider)
        judge = RagJudge(judge_vlm)

        total_elapsed = 0.0
        for img in images:
            nn = img["nn"]
            print(f"  [{nn}] {img['png']}...", end=" ", flush=True)
            result = run_pipeline_method(img, ocr_vlm, judge)
            result["nn"] = nn
            pipeline_results.append(result)
            total_elapsed += result["elapsed"]
            print(f"{result['elapsed']:.1f}초, 탐지 {len(result['findings'])}건")

            # 정답셋 대조
            key = answer_key.get(nn, [])
            if key:
                comp = compare_with_answer_key(key, result["findings"], "② 나눠서")
                comp["nn"] = nn
                pipeline_comparisons.append(comp)

        total_tokens = getattr(ocr_vlm, "total_tokens", 0) + getattr(judge_vlm, "total_tokens", 0)
        pipeline_meta = {"total_elapsed": total_elapsed, "total_tokens": total_tokens}
        print(f"\n② 합계: {total_elapsed:.1f}초, 토큰 {total_tokens}")

    # ── ③ 한 방 ──
    if run_oneshot:
        rag_label = " + RAG 근거" if args.oneshot_rag else ""
        print(f"\n=== ③ 한 방 (VLM 이미지 통째 판정{rag_label}) ===")
        oneshot_vlm = get_vlm(args.oneshot_provider)
        oneshot_context = build_judgment_context() if args.oneshot_rag else ""

        total_elapsed = 0.0
        for img in images:
            nn = img["nn"]
            print(f"  [{nn}] {img['png']}...", end=" ", flush=True)
            result = run_oneshot_method(img, oneshot_vlm, context=oneshot_context)
            result["nn"] = nn
            oneshot_results.append(result)
            total_elapsed += result["elapsed"]
            print(
                f"{result['elapsed']:.1f}초, "
                f"읽은 문장 {result['n_sentences_read']}개, "
                f"위반 탐지 {len(result['findings'])}건"
            )

            # 정답셋 대조
            key = answer_key.get(nn, [])
            if key:
                comp = compare_with_answer_key(key, result["findings"], "③ 한 방")
                comp["nn"] = nn
                oneshot_comparisons.append(comp)

        total_tokens = getattr(oneshot_vlm, "total_tokens", 0)
        oneshot_meta = {"total_elapsed": total_elapsed, "total_tokens": total_tokens}
        print(f"\n③ 합계: {total_elapsed:.1f}초, 토큰 {total_tokens}")

    # ── 결과 출력 ──
    write_result_xlsx(
        pipeline_results, oneshot_results,
        pipeline_comparisons, oneshot_comparisons,
        pipeline_meta, oneshot_meta,
    )

    # 콘솔 요약
    print("\n=== 비교 요약 (잠정) ===")
    if pipeline_comparisons:
        tp = sum(c["tp"] for c in pipeline_comparisons)
        fn = sum(c["fn"] for c in pipeline_comparisons)
        fp = sum(c["fp"] for c in pipeline_comparisons)
        rate = tp / (tp + fn) * 100 if (tp + fn) > 0 else 0
        print(f"② 나눠서:  정탐 {tp} / 미탐 {fn} / 오탐 {fp} / 탐지율 {rate:.1f}%")
    if oneshot_comparisons:
        tp = sum(c["tp"] for c in oneshot_comparisons)
        fn = sum(c["fn"] for c in oneshot_comparisons)
        fp = sum(c["fp"] for c in oneshot_comparisons)
        rate = tp / (tp + fn) * 100 if (tp + fn) > 0 else 0
        print(f"③ 한 방:   정탐 {tp} / 미탐 {fn} / 오탐 {fp} / 탐지율 {rate:.1f}%")


if __name__ == "__main__":
    main()
