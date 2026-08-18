# -*- coding: utf-8 -*-
"""전성분 반영 여부로 나눠, 저장된 두 파이프라인 결과(xlsx)의 격상율 변화를 비교한다.

정보부족형·위반의심형 격상율이 §2-1-3 -> 이번 사이에 움직인 게 "전성분 효과"인지
"실행 편차"인지 가르려고 만들었다(하니 지적, 2026-08-18). §2-1-5에서 이미 손으로
확인한 사실 — 내려간 12건 중 8건만 전성분 반영 이미지에서 나오고 나머지 4건은
전성분과 무관한 이미지에서 나왔다 — 을 지표 자체로 드러낸다.

전성분이 실제로 반영된 이미지(그룹A)와 안 된 이미지(그룹B, 코드 경로가 §2-1-3과
완전히 동일)로 나눠서 같은 지표를 따로 본다. **그룹B는 무료 대조군이다** — 코드가
안 바뀌었는데 결과가 바뀐 만큼이 그대로 실행 편차 크기다. 반복 실행 없이도 그룹A의
변화가 노이즈보다 큰지 가늠할 수 있다.

VLM을 안 부른다. 이미 저장된 결과 xlsx 두 개와 정답셋만 읽는다(API 비용 0).

사용법(backend/에서):
  python scripts/ingredients_group_diff.py \
    --baseline 11st_probe_cosmetic/read_test/ocr_comparison_result_combined_pipeline_oneshot.xlsx \
    --current 11st_probe_cosmetic/read_test/ocr_comparison_result.xlsx
"""
import argparse
import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, "src")
sys.path.insert(0, "scripts")
import compare_ocr  # noqa: E402

_LABEL_XLSX = Path("11st_probe_cosmetic/read_test/label_worksheet_combined.xlsx")
_INGREDIENTS_MAP = Path("11st_probe_cosmetic/read_test/ingredients_map.json")

_KINDS = (
    ("위반", lambda r: r["label"] == "위반"),
    ("정보부족형", lambda r: r["review_kind"] == "정보부족형"),
    ("위반의심형", lambda r: r["review_kind"] == "위반의심형"),
)


def review_kind_by_key(label_xlsx: Path) -> dict[tuple[str, str], str]:
    """(이미지, 문장) -> review_kind(정보부족형/위반의심형/''). 위반 라벨은 ''."""
    key = compare_ocr.load_answer_key(label_xlsx=label_xlsx)
    return {
        (nn, row["sentence"]): row.get("review_kind", "")
        for nn, rows in key.items()
        for row in rows
    }


def load_ingredients_nns(path: Path) -> set[str]:
    """전성분이 실제로 확보된 이미지번호 집합(그룹A)."""
    data = json.loads(path.read_text(encoding="utf-8"))
    return {nn for nn, v in data.items() if v.get("ingredients")}


def load_pipeline_outcomes(result_xlsx: Path, review_kind: dict[tuple[str, str], str]) -> list[dict]:
    """'상세 대조 (잠정)' 시트에서 ② 나눠서 행만, review_kind를 붙여서 낸다.

    같은 (이미지, 문장)이 시트에 **서로 다른 의미로 두 번** 나올 수 있다: 정답셋
    기준 정탐/미탐 판정(사람 판정=검토필요/위반)과, 시스템 finding이 우연히 같은
    텍스트라 오탐으로 잡힌 판정(사람 판정=합법/대상외)이 별개 루프에서 각각
    기록되기 때문이다(2026-08-18 실측, 이미지26 "HAS2..." 사례로 발견). 그래서
    review_kind는 사람 판정이 **검토필요인 행에만** 붙인다 — 안 그러면 우연히
    문장이 겹친 오탐 행까지 정보부족형/위반의심형 집계에 섞여 들어간다.
    """
    wb = openpyxl.load_workbook(result_xlsx)
    ws = wb["상세 대조 (잠정)"]
    rows = []
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 1).value) != "②":
            continue
        nn = str(ws.cell(r, 2).value or "").strip()
        sentence = str(ws.cell(r, 3).value or "").strip()
        label = str(ws.cell(r, 4).value or "").strip()
        result = str(ws.cell(r, 7).value or "")
        rows.append({
            "nn": nn,
            "sentence": sentence,
            "label": label,
            "hit": "정탐" in result,
            "review_kind": review_kind.get((nn, sentence), "") if label == "검토필요" else "",
        })
    return rows


def group_rate(rows: list[dict], kind_filter, in_group) -> tuple[int, int] | None:
    """(정탐, 전체) 또는 해당 문장이 없으면 None."""
    subset = [r for r in rows if kind_filter(r) and in_group(r)]
    if not subset:
        return None
    return sum(1 for r in subset if r["hit"]), len(subset)


def _in_group(nns: set[str], invert: bool = False):
    """`r["nn"] in nns`(또는 반대)를 group_rate 필터로 쓰기 위한 얇은 함수 팩토리."""
    return (lambda r: r["nn"] not in nns) if invert else (lambda r: r["nn"] in nns)


def print_report(before: list[dict], after: list[dict], ing_nns: set[str]) -> None:
    group_a = _in_group(ing_nns)
    group_b = _in_group(ing_nns, invert=True)

    for glabel, gfilter in (("그룹A: 전성분 반영", group_a), ("그룹B: 전성분 미반영(대조군, 코드 동일)", group_b)):
        print(f"\n=== {glabel} ===")
        for klabel, kfilter in _KINDS:
            b = group_rate(before, kfilter, gfilter)
            a = group_rate(after, kfilter, gfilter)
            if b is None or a is None:
                print(f"  {klabel:8} (해당 문장 없음)")
                continue
            btp, btot = b
            atp, atot = a
            brate, arate = (btp / btot * 100), (atp / atot * 100)
            print(f"  {klabel:8} {btp}/{btot}={brate:5.1f}%  ->  {atp}/{atot}={arate:5.1f}%  ({arate - brate:+.1f}%p)")

    print("\n=== 해석 보조: 그룹B(대조군) 변화폭 = 실행 편차 크기 추정 ===")
    for klabel, kfilter in _KINDS:
        b = group_rate(before, kfilter, group_b)
        a = group_rate(after, kfilter, group_b)
        if not b or not a or not b[1] or not a[1]:
            continue
        noise = abs((a[0] / a[1] - b[0] / b[1]) * 100)
        print(f"  {klabel:8} 대조군 변화폭 ≈ {noise:.1f}%p  (그룹A 변화가 이보다 커야 신호로 볼 만하다)")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--baseline", required=True, help="§2-1-3 등 이전 결과 xlsx")
    ap.add_argument("--current", required=True, help="이번 결과 xlsx")
    ap.add_argument("--label-file", default=str(_LABEL_XLSX))
    ap.add_argument("--ingredients-map", default=str(_INGREDIENTS_MAP))
    args = ap.parse_args()

    review_kind = review_kind_by_key(Path(args.label_file))
    ing_nns = load_ingredients_nns(Path(args.ingredients_map))
    before = load_pipeline_outcomes(Path(args.baseline), review_kind)
    after = load_pipeline_outcomes(Path(args.current), review_kind)

    print(f"전성분 반영 이미지(그룹A): {len(ing_nns)}장 — {sorted(ing_nns)}")
    print_report(before, after, ing_nns)


if __name__ == "__main__":
    main()
