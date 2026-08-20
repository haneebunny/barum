# -*- coding: utf-8 -*-
"""평가 골드셋 ver2 생성 — ver1(유형 1축)을 현행 정답셋 기준 2축으로 갱신.

## 왜 필요한가

ver1(`data/cosmetic_eval_labeling.xlsx`)은 **위반유형 1축**만 갖고 있다
(합법/1호/2호/5호/대상외/애매). "검토필요"를 적을 칸 자체가 없다 — 검토필요 정책
(PR #175) 이전에 만들어졌기 때문이다. 그래서 검토필요여야 할 문장을 전부 합법
아니면 위반유형 중 하나로 밀어넣을 수밖에 없었다.

그 결과 ver1로 잰 지표를 못 믿게 됐다. 2026-08-20 실측: ver1과 현행 963문장
정답셋(`label_worksheet_combined.xlsx`)에 겹치는 35문장 중 **14건(40%)이 합법 여부부터
서로 다르다**. `eval_ragjudge.py`가 보고한 "오탐 12건" 중 최소 7건은 AI가 현행
정답셋과 일치하는 판정을 냈는데 ver1의 낡은 라벨 때문에 오탐으로 집계된 허수였다.
이 지표로 프롬프트 A/B를 돌리면 낡은 라벨 쪽으로 최적화된다.

**963문장 정답셋은 낡지 않았다. 낡은 건 이 40문장 평가셋이다.**
(상세: `docs/result/2026-08-20_판정로직_고도화_로그.md` ⑥)

## 무엇을 하는가

1. ver1의 43행을 읽는다.
2. 현행 963셋에 같은 문장이 있으면(35건) **그 라벨을 그대로 이식**한다. 이미 검수를
   거친 라벨이라 사람이 다시 볼 필요가 없다.
3. 963셋에 없는 8건은 아래 `_MANUAL_V2`의 값을 쓴다(팀장 승인 2026-08-20).
4. `data/cosmetic_eval_labeling_v2.xlsx`로 쓴다. **ver1은 안 건드린다**(예전 기록
   삭제 금지 — 과거 측정치와의 연속성을 위해 원본을 남긴다).

`backend/data/`는 gitignore라 xlsx 자체는 커밋되지 않는다. 그래서 라벨 결정을 이
스크립트에 담아 리뷰·재현이 가능하게 했다. 새 환경에서는 이걸 한 번 돌리면 된다.

    ./venv/bin/python scripts/build_eval_goldset_v2.py

## ver2 스키마 (2축)

| 열 | 값 | 뜻 |
|---|---|---|
| 라벨 | 합법·1호·2호·5호·대상외·애매 | 위반 **유형** |
| 확정도 | 위반·검토필요·(빈칸) | 위반 **확정도**. 합법·대상외·애매는 빈칸 |

963셋은 검토필요 160건 중 154건이 위반유형 칸이 비어 있다. 검토필요는 "유형은
정해도 확정을 못 하는" 상태라 유형까지 매기지 않은 것이다. 그래서 ver2도 그런 행은
유형을 비워 두고, 채점기가 그런 행은 **확정도 축만** 보게 한다(유형 불문 플래그
여부만 확인). 위반 확정분은 유형까지 맞춰야 정답이다.
"""

import shutil
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, "scripts")
import compare_ocr  # noqa: E402

_V1 = Path("data/cosmetic_eval_labeling.xlsx")
_V2 = Path("data/cosmetic_eval_labeling_v2.xlsx")
_ANSWER_KEY = Path("11st_probe_cosmetic/read_test/label_worksheet_combined.xlsx")

# 963셋에 없어 이식할 라벨이 없는 8건. 팀장 승인(2026-08-20).
# 값은 (라벨, 확정도, 판단근거). 확정도는 합법·대상외·애매면 "".
_MANUAL_V2: dict[str, tuple[str, str, str]] = {
    "수분을 채워야 하니깐, 물부터 남다르게!": (
        "합법", "", "일반 보습 소구 + 감성 카피. 효능 표방 없음(ver1 판단 유지).",
    ),
    "눈가, 주름고민 모두 다 해결하는 안나홀츠 안티링클 아이크림": (
        "2호_기능성오인", "검토필요",
        "주름개선 표방이라 2호. 심사·보고 여부를 알 수 없어 확정 불가.",
    ),
    "효과로 증명합니다.": (
        "5호_거짓과장기만", "검토필요",
        "근거 없는 효과 단정. 다만 PR#152에서 '구체적 검증방법 언급이 없어 "
        "근거없는검증방법 규칙 대상은 아님'으로 확인됐다 — 규칙이 아니라 VLM이 "
        "판단할 영역이라 검토필요로 둔다.",
    ),
    "진짜예요!!": (
        "5호_거짓과장기만", "검토필요",
        "바로 앞 문장('효과로 증명합니다.')에 이어지는 감탄사. **문장 단독으로는 "
        "효능어도 주장도 없어 문맥 없이는 판정 불가** — 문장 독립 채점 방식에선 "
        "구조적으로 못 잡는다. 라벨은 사람이 문맥 보고 매긴 값을 유지하되, "
        "이 행의 미탐은 모델 결함으로 읽지 말 것(채점 시 캐비어트).",
    ),
    "Holtz for men": (
        "대상외", "", "브랜드명 표기. 광고 문구가 아니라 판정 대상 자체가 아니다.",
    ),
    "피부 내 콜라겐과 엘라스틴 생성량을 증가시켜 주름개선에 도움을 줍니다.": (
        "2호_기능성오인", "검토필요",
        "주름개선 표방이라 2호. 콜라겐 증가는 PR#167에서 violation -> needs_review로 "
        "옮겼다(prohibited_expressions.md §3 실증대상, 위반 단정 금지).",
    ),
    "무겁지 않은 제형으로 피부에 충분한 영양과 보습을 공급하고, 피부 탄력을 높이며 주름 개선 및 모공쪼임에 도움을 주는 기능성 화장품": (
        "2호_기능성오인", "검토필요",
        "'주름 개선'·'기능성 화장품' 표방이라 2호. ver1은 합법이었는데 검토필요 정책 "
        "이전 라벨이다. 모공쪼임은 별건(2026-08-19 규칙 논의에서 라벨 '애매'로 제외).",
    ),
    "미백• 주름개선 이중기능성 화장품": (
        "2호_기능성오인", "검토필요",
        "기능성 표방. 인정문구(approved_efficacy_statements.md §1~§3)와는 다른 "
        "축약 표기라 인정문구 예외 대상이 아니다. 심사·보고 여부 확인 필요.",
    ),
}

# 963셋 판정 -> (ver2 라벨 기본값, ver2 확정도). 유형이 963셋에 있으면 라벨을 그걸로 덮는다.
_JUDGMENT_MAP: dict[str, tuple[str, str]] = {
    "합법": ("합법", ""),
    "대상외": ("대상외", ""),
    "애매": ("애매", ""),
    "위반": ("", "위반"),          # 라벨은 위반유형 칸에서 채운다
    "검토필요": ("", "검토필요"),   # 유형 칸이 비어 있으면 그대로 빈칸(유형 불문 채점)
}


def _load_answer_key() -> dict[str, tuple[str, str]]:
    """현행 963 정답셋을 {문장: (판정, 위반유형)}으로 읽는다."""
    key = compare_ocr.load_answer_key(label_xlsx=_ANSWER_KEY)
    out: dict[str, tuple[str, str]] = {}
    for rows in key.values():
        for row in rows:
            sentence = (row["sentence"] or "").strip()
            if sentence:
                out[sentence] = (row["judgment"], row.get("violation_type") or "")
    return out


def build() -> None:
    """ver1 + 963셋 + 수동 8건을 합쳐 ver2 xlsx를 쓴다."""
    if not _V1.exists():
        sys.exit(f"[없음] {_V1} — ver1이 있어야 ver2를 만든다")
    answer = _load_answer_key()

    shutil.copy(_V1, _V2)  # 서식·출처열을 그대로 물려받고 라벨 열만 갈아끼운다
    wb = openpyxl.load_workbook(_V2)
    ws = wb["라벨링"]

    # 확정도 열을 라벨 열 바로 뒤에 끼운다(라벨=4열 고정, ver1 스키마).
    ws.insert_cols(5)
    ws.cell(1, 5).value = "확정도"
    last = ws.max_column + 1
    ws.cell(1, last).value = "ver2_출처"

    n_ported = n_manual = n_changed = 0
    for r in range(2, ws.max_row + 1):
        sentence = str(ws.cell(r, 3).value or "").strip()
        old_label = str(ws.cell(r, 4).value or "").strip()
        if not sentence or not old_label:
            continue

        if sentence in answer:
            judgment, vtype = answer[sentence]
            label, flag = _JUDGMENT_MAP.get(judgment, (judgment, ""))
            if not label:
                label = vtype  # 위반·검토필요는 유형 칸에서(비어 있으면 유형 불문)
            source = "963정답셋 이식"
            n_ported += 1
        elif sentence in _MANUAL_V2:
            label, flag, reason = _MANUAL_V2[sentence]
            source = f"수동(팀장 승인 2026-08-20): {reason}"
            n_manual += 1
        else:
            # ver1에만 있고 두 경로 어디에도 없는 문장. 라벨을 지어내지 않는다.
            ws.cell(r, last).value = "미갱신(963셋·수동목록 어디에도 없음)"
            continue

        if label != old_label:
            n_changed += 1
        ws.cell(r, 4).value = label
        ws.cell(r, 5).value = flag
        ws.cell(r, last).value = source

    wb.save(_V2)
    print(f"ver2 저장: {_V2}")
    print(f"  963셋에서 이식: {n_ported}건")
    print(f"  수동 라벨(승인분): {n_manual}건")
    print(f"  ver1 대비 유형라벨이 바뀐 행: {n_changed}건")
    print(f"  (ver1 {_V1}은 그대로 보존)")


if __name__ == "__main__":
    build()
