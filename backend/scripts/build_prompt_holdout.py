# -*- coding: utf-8 -*-
"""프롬프트 A/B용 홀드아웃 셋을 만든다 (963셋 중 평가셋과 안 겹치는 문장에서 층화표본).

## 왜 필요한가

프롬프트 A/B를 42문장 ver2 골드셋으로만 재고 있었다. 두 가지 문제가 있다.

1. **표본이 너무 작다.** 실행 편차가 42문항에서 ±2~3건(약 ±7%p)이라, 그보다 작은
   효과는 **검출 자체가 안 된다.** CoT A/B에서 A 32~35 / B 33~34가 나왔는데 이건
   "효과 없음"이 아니라 **"이 표본으로는 못 가린다"**는 뜻이다. 둘은 다르다.
2. **선택 편향.** 같은 42문장으로 여러 프롬프트 안을 반복 비교하면 그 셋에 잘 맞는
   안을 고르게 된다(PM 지적, 2026-08-20).

## 이 셋이 홀드아웃인 범위 (한계를 분명히 한다)

963문장 정답셋 중 **ver2 평가셋(42문장)과 안 겹치는 956건**에서 뽑는다. 라벨이 이미
검수돼 있어 추가 라벨링 비용이 0이다.

- **프롬프트 A/B에 대해서는 홀드아웃이다.** 지금까지 프롬프트 실험은 전부 42문장
  골드셋으로만 했고 이 956건은 한 번도 안 썼다.
- **규칙 변경에 대해서는 홀드아웃이 아니다.** `rule_sweep`이 963문장 전체를 훑어 왔으므로
  규칙은 이 문장들을 이미 봤다. 규칙 A/B에 이 셋을 쓰면 안 된다.

## 층화

정답 분포가 심하게 치우쳐 있다(대상외 617 / 합법 178 / 검토필요 147 / 위반 12).
그대로 무작위로 뽑으면 표본 대부분이 대상외가 되어 판정 성능을 못 잰다. 그래서
**판정이 필요한 라벨(위반·검토필요)을 우선 채우고, 합법·대상외를 대조군으로 섞는다.**
비율은 ver2 골드셋과 비슷하게 맞춘다(위반·검토필요가 절반 이상).

    ./venv/bin/python scripts/build_prompt_holdout.py --size 120
"""

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, "scripts")
import compare_ocr  # noqa: E402

_ANSWER_KEY = Path("11st_probe_cosmetic/read_test/label_worksheet_combined.xlsx")
_EVAL_V2 = Path("data/cosmetic_eval_labeling_v2.xlsx")
_OUT = Path("data/prompt_holdout.jsonl")

# 층화 비율(판정이 필요한 라벨을 우선). 합계가 1.0이 아니어도 되며, 부족하면 있는 만큼만.
_QUOTA = {"위반": 0.10, "검토필요": 0.45, "합법": 0.25, "대상외": 0.20}


def _eval_sentences() -> set[str]:
    """ver2 평가셋 문장(홀드아웃에서 빼야 할 것)."""
    if not _EVAL_V2.exists():
        sys.exit(f"[없음] {_EVAL_V2} — build_eval_goldset_v2.py를 먼저 돌릴 것")
    wb = openpyxl.load_workbook(_EVAL_V2)
    ws = wb["라벨링"]
    return {
        str(ws.cell(r, 3).value or "").strip()
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, 3).value
    }


def build(size: int) -> None:
    """층화표본을 뽑아 jsonl로 쓴다. 무작위를 안 쓴다(재현 가능해야 하므로)."""
    excluded = _eval_sentences()
    key = compare_ocr.load_answer_key(label_xlsx=_ANSWER_KEY)

    # 라벨별로 모으되, 같은 문장이 여러 이미지에 중복 등장하므로 첫 것만 쓴다.
    by_label: dict[str, list[dict]] = defaultdict(list)
    seen: set[str] = set()
    for rows in key.values():
        for row in rows:
            s = (row["sentence"] or "").strip()
            if not s or s in excluded or s in seen:
                continue
            seen.add(s)
            by_label[row["judgment"]].append(
                {"text": s, "human": row.get("violation_type") or "",
                 "flag": row["judgment"]}
            )

    print(f"후보 풀(평가셋 제외, 중복 제거): {sum(len(v) for v in by_label.values())}건")
    for lab, items in sorted(by_label.items()):
        print(f"  {lab}: {len(items)}건")

    out: list[dict] = []
    for label, ratio in _QUOTA.items():
        want = round(size * ratio)
        pool = by_label.get(label, [])
        # 무작위 대신 앞에서부터 균등 간격으로 뽑는다 — seed 없이도 재현된다.
        if not pool:
            continue
        step = max(1, len(pool) // want) if want else 1
        picked = pool[::step][:want]
        for p in picked:
            # 채점기 스키마에 맞춘다: 합법·대상외·애매는 확정도 빈칸.
            flag = p["flag"] if p["flag"] in ("위반", "검토필요") else ""
            human = p["human"] if p["human"] else (
                p["flag"] if p["flag"] in ("합법", "대상외") else ""
            )
            out.append({"n": len(out) + 1, "text": p["text"],
                        "human": human, "flag": flag})
        print(f"  -> {label}: {len(picked)}건 선정(목표 {want})")

    _OUT.write_text(
        "\n".join(json.dumps(o, ensure_ascii=False) for o in out) + "\n",
        encoding="utf-8",
    )
    print(f"\n홀드아웃 저장: {_OUT} ({len(out)}건)")
    print("⚠ 프롬프트 A/B 전용이다. 규칙 A/B에는 쓰면 안 된다(rule_sweep이 이미 본 문장들).")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--size", type=int, default=120, help="표본 크기(기본 120)")
    build(ap.parse_args().size)
