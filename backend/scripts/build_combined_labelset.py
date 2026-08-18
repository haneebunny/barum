"""§2 3-way 재측정용: 기존 240문장 + 확장 835문장 = 1,075문장 정답셋 통합.

원본(label_worksheet_reviewed.xlsx·label_worksheet_expansion.xlsx)은 감사추적용
으로 그대로 두고, compare_ocr.py가 바로 읽을 수 있는 "라벨링" 시트 스키마
(이미지/상품코드/문장#/문장/판정/위반유형)로 통합 파일을 새로 만든다.

최종 라벨 우선순위(PM 2026-08-17 확정):
- 기존 240문장(label_worksheet_reviewed.xlsx): 비비_최종판단(L열, 불일치 36건
  해소) > 판정(E열, 대수 원안 — LLM과 일치했던 204건). compare_ocr.py의
  load_answer_key()가 이미 이 로직을 그대로 구현하고 있어 재사용한다.
- 신규 835문장(label_worksheet_expansion.xlsx): 합의판단(O열, 하니 최종, 최우선)
  > 비비_최종판단(N열, 인용검증) > 확정_판정(L열, 두 모델 자동일치).
  합의판단은 "라벨 — 근거" 형식이 아니라 자유서술 10건이라 자동파싱 대신
  하니 확인을 거친 수동 매핑(_HANI_OVERRIDES)을 쓴다.

이미지번호(nn)는 기존 01~14, 신규 15~53으로 이미 안 겹친다(상품코드도 겹침
없음, 2026-08-17 확인).

실행(backend/에서):
    python scripts/build_combined_labelset.py

출력:
    11st_probe_cosmetic/read_test/label_worksheet_combined.xlsx  (라벨링 시트, 1075행)
    11st_probe_cosmetic/read_test/_combined_answer_key.json      (nn -> png, 53장)
"""
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side

_READ_TEST = Path("11st_probe_cosmetic/read_test")
_REVIEWED_XLSX = _READ_TEST / "label_worksheet_reviewed.xlsx"
_EXPANSION_XLSX = _READ_TEST / "label_worksheet_expansion.xlsx"
_ORIG_ANSWER_KEY = _READ_TEST / "_llm_answer_key.json"
_EXPANSION_ANSWER_KEY = _READ_TEST / "_expansion_answer_key.json"
_OUT_XLSX = _READ_TEST / "label_worksheet_combined.xlsx"
_OUT_ANSWER_KEY = _READ_TEST / "_combined_answer_key.json"

_LABELS = {"합법", "1호_의약품오인", "2호_기능성오인", "5호_거짓과장기만", "대상외", "위반", "검토필요"}


def _parse_prefixed_label(text: str) -> str | None:
    """"검토필요 — 근거설명..." 형식에서 라벨만 뽑는다. 규격 밖이면 None."""
    label = text.split("—", 1)[0].split("-", 1)[0].strip()
    return label if label in _LABELS else None


# 확장셋 "합의판단"(O열) 10건은 자유서술이라 자동파싱이 안 된다. 문장·모델판정·
# 비비판단까지 다 읽고 하니에게 확인받은 수동 매핑(2026-08-17). row는 라벨링
# 시트의 엑셀 행번호(1-based, 헤더 포함) — label_worksheet_expansion.xlsx 기준.
_HANI_OVERRIDES: dict[int, tuple[str, str]] = {
    419: ("위반", "5호_거짓과장기만"),  # "거짓 없는 화장품, 효과로 증명하는" — type_5 예시
    476: ("위반", "5호_거짓과장기만"),  # 위와 동일 문구(다른 이미지)
    556: ("위반", "5호_거짓과장기만"),  # 위조 증빙 서류 재사용(비비 발견, 하니 확인)
    588: ("위반", "1호_의약품오인"),     # "치유하는 힘" — 치료/치유/힐링 원칙
    598: ("위반", "1호_의약품오인"),     # "치료하는 연고의 주성분"
    642: ("위반", "5호_거짓과장기만"),  # 556과 같은 위조 증빙 재사용
    737: ("합법", ""),                  # "시카 엑소좀" — 원료 대분류 예외(엑소좀 버그와 동일 사례)
    743: ("위반", "5호_거짓과장기만"),  # "Pin" 침투 메커니즘(니들류, T5 매핑 정정)
    744: ("위반", "5호_거짓과장기만"),  # "15㎛ Pin"
    751: ("위반", "5호_거짓과장기만"),  # "미세침이... 침투"
}


def build_expansion_rows() -> list[dict]:
    """label_worksheet_expansion.xlsx에서 (nn, code, seq, sentence, label, vtype) 835건."""
    wb = openpyxl.load_workbook(_EXPANSION_XLSX)
    ws = wb["라벨링"]
    rows = []
    n_hani = n_bibi = n_auto = n_unresolved = 0
    for r in range(2, ws.max_row + 1):
        nn = str(ws.cell(r, 1).value or "").strip()
        if not nn:
            continue
        sentence = str(ws.cell(r, 4).value or "").strip()

        if r in _HANI_OVERRIDES:
            label, vtype = _HANI_OVERRIDES[r]
            n_hani += 1
        else:
            bibi_raw = ws.cell(r, 14).value
            if bibi_raw:
                label = _parse_prefixed_label(str(bibi_raw))
                if label is None:
                    print(f"    [경고] row {r}: 비비_최종판단 파싱 실패 '{bibi_raw}', 확정_판정으로 폴백")
                    label = str(ws.cell(r, 12).value or "").strip()
                    vtype = str(ws.cell(r, 13).value or "").strip()
                else:
                    n_bibi += 1
                    # 위반유형은 비비 근거 자유텍스트라 구조화 안 함 — 확정_위반유형(두 모델
                    # 자동일치분 기준)으로 채운다. 채점(위반/검토필요 여부)엔 안 쓰이는
                    # 정보용 필드라 부정확해도 §2 결과엔 영향 없다.
                    vtype = str(ws.cell(r, 13).value or "").strip()
            else:
                label = str(ws.cell(r, 12).value or "").strip()
                vtype = str(ws.cell(r, 13).value or "").strip()
                if label:
                    n_auto += 1
                else:
                    n_unresolved += 1
                    continue  # 자동확정도 안 됐고 비비/하니도 안 본 행 — 정답셋에서 제외

        rows.append({
            "nn": nn, "code": str(ws.cell(r, 2).value or "").strip(), "seq": ws.cell(r, 3).value,
            "sentence": sentence, "label": label, "vtype": vtype,
        })
    print(f"  확장셋: 하니override {n_hani} / 비비판단 {n_bibi} / 자동확정 {n_auto} / 미해소제외 {n_unresolved}")
    return rows


def build_original_rows() -> list[dict]:
    """label_worksheet_reviewed.xlsx에서 240건 — compare_ocr.py의 load_answer_key()와 같은 우선순위."""
    wb = openpyxl.load_workbook(_REVIEWED_XLSX)
    ws = wb["라벨링"]
    rows = []
    for r in range(2, ws.max_row + 1):
        nn = str(ws.cell(r, 1).value or "").strip()
        sentence = str(ws.cell(r, 4).value or "").strip()
        if not nn or not sentence:
            continue
        label = str(ws.cell(r, 5).value or "").strip()
        vtype = str(ws.cell(r, 6).value or "").strip()
        final = str(ws.cell(r, 12).value or "").strip()
        if final:
            parsed = final.split("—", 1)[0].strip()
            if parsed in _LABELS:
                label = parsed
                vtype = next((t for t in ("1호_의약품오인", "2호_기능성오인", "5호_거짓과장기만") if t in final), vtype)
        rows.append({
            "nn": nn, "code": str(ws.cell(r, 2).value or "").strip(), "seq": ws.cell(r, 3).value,
            "sentence": sentence, "label": label, "vtype": vtype,
        })
    return rows


def write_combined_xlsx(orig_rows: list[dict], expansion_rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "라벨링"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin = Border(*(Side(style="thin"),) * 4)

    headers = ["이미지", "상품코드", "문장#", "문장", "판정", "위반유형"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.border = header_font, header_fill, thin
    widths = [8, 12, 8, 50, 10, 16]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    row = 2
    for r in orig_rows + expansion_rows:
        vals = [r["nn"], r["code"], r["seq"], r["sentence"], r["label"], r["vtype"]]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=row, column=ci, value=v).border = thin
        row += 1

    ws.auto_filter.ref = f"A1:F{row - 1}"
    ws.freeze_panes = "A2"
    wb.save(_OUT_XLSX)
    print(f"저장: {_OUT_XLSX} (총 {row - 2}행)")


def write_combined_answer_key() -> None:
    orig = json.loads(_ORIG_ANSWER_KEY.read_text(encoding="utf-8"))
    expansion = json.loads(_EXPANSION_ANSWER_KEY.read_text(encoding="utf-8"))
    combined = [{"nn": m["nn"], "png": m["png"]} for m in orig] + \
               [{"nn": m["nn"], "png": m["png"]} for m in expansion]
    _OUT_ANSWER_KEY.write_text(json.dumps(combined, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"저장: {_OUT_ANSWER_KEY} (이미지 {len(combined)}장)")


def main() -> None:
    print("정답셋 통합 중...")
    orig_rows = build_original_rows()
    expansion_rows = build_expansion_rows()
    print(f"  기존 {len(orig_rows)}건 + 신규 {len(expansion_rows)}건 = {len(orig_rows) + len(expansion_rows)}건")

    from collections import Counter
    dist = Counter(r["label"] for r in orig_rows + expansion_rows)
    print(f"  라벨 분포: {dict(dist)}")
    n_violation = dist.get("위반", 0) + dist.get("검토필요", 0)
    print(f"  위반+검토필요 합계: {n_violation}")

    write_combined_xlsx(orig_rows, expansion_rows)
    write_combined_answer_key()


if __name__ == "__main__":
    main()
