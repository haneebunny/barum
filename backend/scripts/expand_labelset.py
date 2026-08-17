"""평가 표본 확장 — 신규 상품 OCR 추출 + 서로 다른 AI 모델 2개 독립 초안.

기존 240문장 정답셋(위반+검토필요 23건)은 표본이 작아 1건만 뒤집혀도 탐지율이
4.3%p씩 흔들린다(§2-1-2 실측). 위반+검토필요 50건+ 확보를 목표로 신규 상품
21개(기존 8개와 중복 없음, `collect_11st_details.py`로 실전형 분포 — 기능성
위반 위주가 아니라 합법/대상외 비중도 섞어서 수집)에서 이미지 39장을 골랐다.

이번엔 대수의 "사람 1차 판정"이 없다. 대신 **서로 다른 AI 모델 2개가 독립적으로
초안을 낸다**(review_labelset.py와 같은 원칙 — 둘 다 판정 프로덕션 모델
gpt-5-mini와 달라야 순환참조가 안 생긴다). 두 모델이 일치하면 그대로 확정,
불일치만 비비가 인용 검증한다(이미 확정된 3축 규칙을 기계적으로 적용 가능).

실행(backend/에서):
    python scripts/expand_labelset.py --ocr              # ① 신규 이미지 OCR 문장 추출
    python scripts/expand_labelset.py --draft             # ② 모델 2개 독립 초안(①끝난 뒤)
    python scripts/expand_labelset.py --dry                # 호출 없이 배선만 점검
    python scripts/expand_labelset.py --draft --limit 2   # 이미지 2장만(스모크 테스트)

출력:
    11st_probe_cosmetic/read_test/_expansion_answer_key.json   (① OCR 결과)
    11st_probe_cosmetic/read_test/label_worksheet_expansion.xlsx (② 듀얼모델 초안)
"""
import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, "src")
sys.path.insert(0, ".")  # tile_split.py (backend 루트)
from barum.models import JudgmentFlag, ViolationType  # noqa: E402
from barum.pipeline import _ocr_image  # noqa: E402
from barum.reference.context import build_judgment_context  # noqa: E402
from barum.vlm import get_vlm  # noqa: E402

_READ_TEST = Path("11st_probe_cosmetic/read_test")
_IMAGES_DIR = _READ_TEST / "images"
_EXPANSION_KEY = _READ_TEST / "_expansion_answer_key.json"
_OUT_XLSX = _READ_TEST / "label_worksheet_expansion.xlsx"

# 신규 이미지 범위(nn). 기존 read_test(01~14)와 안 겹치게 15부터 시작.
_NN_RANGE = range(15, 54)

# 프로덕션(gpt-5-mini)과 둘 다 달라야 순환참조가 안 생긴다. luna는 review_labelset.py와
# 같은 모델(재사용). 둘째 모델은 원래 sol(형제 모델)을 쓰려 했는데, sol이 플래그십
# 가격대($5/$30)라 이 물량(835문장×2)이면 그것만 만 원대 중반까지 간다. gpt-5
# ($0.625/$5)로 바꾸면 순환참조 회피 조건(프로덕션 mini와 다른 모델)은 똑같이
# 만족하면서 비용은 2천 원대로 끝난다(PM 2026-08-15 지시).
MODEL_A = "gpt-5.6-luna"
MODEL_B = "gpt-5"

_VIOLATION_TYPES = [
    ViolationType.type_1_drug_misperception.value,
    ViolationType.type_2_functional_misperception.value,
    ViolationType.type_5_deception.value,
]
_LABELS = [ViolationType.legal.value, *_VIOLATION_TYPES, ViolationType.out_of_scope.value, JudgmentFlag.needs_review.value]

REVIEW_PROMPT_TEMPLATE = """너는 한국 화장품 광고 위반 판정 전문가다. 아래 [판정 근거]를 참고해서,
이 이미지에서 이미 추출된 아래 문장들 각각을 판정하라. 문장 자체는 이미 확정됐으니
새로 읽지 말고(이미지에 없는 문장이 섞여 있어도 그대로 판정), 목록 순서 그대로 라벨과 근거만 매겨라.

[판정 근거]
{context}

[문장 목록]
{sentence_list}

라벨(정확히 이 중 하나): 합법 / 1호_의약품오인 / 2호_기능성오인 / 5호_거짓과장기만 / 대상외 / 검토필요
- 미탐(위반을 합법으로 놓침)이 제일 나쁜 실수다. 확신 없으면 검토필요로, 안전 쪽으로 판단하라.
- 대상외는 성분명 나열·브랜드명·인증마크·거래정보 등 광고 문구가 아닌 것.

JSON으로만 답하라:
{{"results": [{{"문장번호": 1, "label": "라벨", "reason": "짧은 근거"}}, ...]}}
문장번호는 반드시 [문장 목록]의 번호와 1:1로 대응해야 하고, 개수도 정확히 같아야 한다."""


def load_expansion_meta() -> list[dict]:
    """이미지 파일명에서 (nn, code, png) 메타를 만든다(collect 스크립트 산출물엔 없어 직접 구성)."""
    entries = []
    for path in sorted(_IMAGES_DIR.glob("*.png")):
        nn_str, rest = path.stem.split("_", 1)
        try:
            nn = int(nn_str)
        except ValueError:
            continue
        if nn not in _NN_RANGE:
            continue
        code = rest.split("_", 1)[0]
        entries.append({"nn": nn_str, "code": code, "png": path.name})
    return entries


# ── ① OCR 문장 추출 ──


def run_ocr(limit: int | None = None) -> None:
    """신규 이미지 39장을 프로덕션 OCR(Gemini, tile_split)로 문장 추출한다."""
    meta = load_expansion_meta()
    if limit:
        meta = meta[:limit]
    print(f"OCR 대상: 이미지 {len(meta)}장")

    vlm = get_vlm("gemini")
    results = []
    for m in meta:
        img_path = _IMAGES_DIR / m["png"]
        print(f"  [{m['nn']}] {m['png']}...", end=" ", flush=True)
        try:
            sentences = _ocr_image(img_path.read_bytes(), m["png"], vlm, verbose=False)
            texts = [s["text"] for s in sentences]
            print(f"{len(texts)}문장")
        except Exception as e:
            print(f"[skip] OCR 실패 {type(e).__name__}: {e}")
            texts = []
        results.append({**m, "llm_sentences": texts})

    _EXPANSION_KEY.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    total_sentences = sum(len(r["llm_sentences"]) for r in results)
    n_empty = sum(1 for r in results if not r["llm_sentences"])
    print(f"\n저장: {_EXPANSION_KEY} (총 문장 {total_sentences}개, 미추출 {n_empty}장)")


# ── ② 듀얼모델 독립 초안 ──


def draft_one_model(nn: str, sentences: list[str], image_bytes: bytes, vlm, context: str) -> dict[int, dict]:
    """review_labelset.py의 review_image()와 같은 방식. 한 모델·한 이미지 판정."""
    numbered = "\n".join(f"{i}. {s}" for i, s in enumerate(sentences, 1))
    prompt = REVIEW_PROMPT_TEMPLATE.format(context=context, sentence_list=numbered)
    try:
        res = vlm.generate_json(prompt, [image_bytes])
    except Exception as e:
        print(f"    [skip] {nn}: 호출 실패 {type(e).__name__}: {e}")
        return {}

    raw = res.get("results", []) if isinstance(res, dict) else []
    if len(raw) != len(sentences):
        print(f"    [skip] {nn}: 응답 {len(raw)}개 != 문장 {len(sentences)}개, 미판정으로 남김")
        return {}

    out = {}
    for item in raw:
        try:
            n = int(item.get("문장번호"))
        except (TypeError, ValueError):
            continue
        label = (item.get("label") or "").strip()
        if label not in _LABELS:
            continue
        out[n] = {"label": label, "reason": (item.get("reason") or "").strip()}
    return out


# 이미지 1장이 문장 80개 넘게 뽑히는 경우(대형 상세페이지)가 있어, 한 번에 다 보내면
# 모델이 개수를 못 맞춰 이미지 전체가 미판정 처리될 위험이 크다(draft_one_model이
# 요청/응답 개수 불일치면 통째로 버림). 청크로 쪼개 위험을 이미지당이 아니라
# 청크당으로 낮춘다.
_MAX_BATCH = 30


def draft_one_model_batched(nn: str, sentences: list[str], image_bytes: bytes, vlm, context: str) -> dict[int, dict]:
    """문장이 많으면 `_MAX_BATCH`개씩 나눠서 draft_one_model을 반복 호출하고 합친다."""
    if len(sentences) <= _MAX_BATCH:
        return draft_one_model(nn, sentences, image_bytes, vlm, context)

    merged: dict[int, dict] = {}
    for offset in range(0, len(sentences), _MAX_BATCH):
        chunk = sentences[offset : offset + _MAX_BATCH]
        chunk_result = draft_one_model(f"{nn}(청크 {offset // _MAX_BATCH + 1})", chunk, image_bytes, vlm, context)
        for local_i, v in chunk_result.items():
            merged[offset + local_i] = v
    return merged


def _label_to_columns(label: str) -> tuple[str, str]:
    if label in _VIOLATION_TYPES:
        return JudgmentFlag.violation.value, label
    return label, ""


def write_expansion_xlsx(meta: list[dict], drafts_a: dict, drafts_b: dict) -> None:
    """두 모델 초안을 나란히 놓고 일치여부를 표시한 워크북을 만든다.

    label_worksheet.xlsx와 같은 컬럼 순서를 유지하되(이미지/상품코드/문장#/문장),
    "판정" 자리에 사람 1차가 없으니 두 모델 컬럼을 나란히 두고, 일치하면 그 값을
    "확정_판정"에 채워 그대로 정답셋으로 쓸 수 있게 한다. 불일치는 빈 채로 둬서
    비비가 인용검증 후 채우게 한다.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "라벨링"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    mismatch_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin = Border(*(Side(style="thin"),) * 4)
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = [
        "이미지", "상품코드", "문장#", "문장",
        f"{MODEL_A}_판정", f"{MODEL_A}_위반유형", f"{MODEL_A}_근거",
        f"{MODEL_B}_판정", f"{MODEL_B}_위반유형", f"{MODEL_B}_근거",
        "일치여부", "확정_판정", "확정_위반유형", "비비_최종판단",
    ]
    widths = [8, 12, 8, 45, 10, 16, 40, 10, 16, 40, 10, 10, 16, 40]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.border = header_font, header_fill, thin
        ws.column_dimensions[c.column_letter].width = w

    row = 2
    n_match = n_mismatch = n_unjudged = 0
    for m in meta:
        nn = m["nn"]
        sentences = m.get("llm_sentences") or []
        a = drafts_a.get(nn, {})
        b = drafts_b.get(nn, {})
        for i, sentence in enumerate(sentences, 1):
            da, db = a.get(i), b.get(i)
            vals = [nn, m["code"], i, sentence]
            if da is None or db is None:
                vals += ["(미판정)", "", "", "(미판정)", "", "", "", "", "", ""]
                n_unjudged += 1
            else:
                la, va = _label_to_columns(da["label"])
                lb, vb = _label_to_columns(db["label"])
                match = la == lb and va == vb
                if match:
                    n_match += 1
                    confirmed = [la, va]
                else:
                    n_mismatch += 1
                    confirmed = ["", ""]
                vals += [la, va, da["reason"], lb, vb, db["reason"],
                         "일치" if match else "불일치", *confirmed, ""]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.border = thin
                if ci in (4, 7, 10):
                    c.alignment = wrap
            if da is not None and db is not None and not match:
                for ci in range(1, len(headers) + 1):
                    ws.cell(row=row, column=ci).fill = mismatch_fill
            row += 1

    ws.auto_filter.ref = f"A1:N{row - 1}"
    ws.freeze_panes = "A2"
    wb.save(_OUT_XLSX)

    total = n_match + n_mismatch + n_unjudged
    print(f"\n저장: {_OUT_XLSX}")
    print(f"일치(자동확정) {n_match} / 불일치(비비 검증 필요) {n_mismatch} / 미판정 {n_unjudged} (총 {total})")


_MISMATCH_XLSX = _READ_TEST / "label_worksheet_expansion_mismatches.xlsx"


def export_mismatches() -> None:
    """label_worksheet_expansion.xlsx에서 불일치 행만 뽑아 비비에게 넘길 워크북을 만든다.

    review_labelset.py의 export_mismatches()와 같은 패턴. 여기선 "대수 1차"가 없으니
    두 모델(A/B) 판정을 나란히 놓는다 — 인용 검증 방식(review_labelset.py에서 확정한
    방식)을 그대로 적용할 수 있게, 어느 쪽이 맞는지 판단 재료(위반유형·근거)를 전부 싣는다.
    """
    wb = openpyxl.load_workbook(_OUT_XLSX)
    ws = wb["라벨링"]

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "불일치"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin = Border(*(Side(style="thin"),) * 4)
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = [
        "이미지", "상품코드", "문장#", "문장",
        f"{MODEL_A}_판정", f"{MODEL_A}_위반유형", f"{MODEL_A}_근거",
        f"{MODEL_B}_판정", f"{MODEL_B}_위반유형", f"{MODEL_B}_근거",
        "비비_최종판단",
    ]
    widths = [8, 12, 8, 45, 10, 16, 40, 10, 16, 40, 40]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = out_ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.border = header_font, header_fill, thin
        out_ws.column_dimensions[c.column_letter].width = w

    row_out = 2
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 11).value or "").strip() != "불일치":
            continue
        values = [ws.cell(r, c).value for c in range(1, 11)] + [""]
        for ci, v in enumerate(values, 1):
            c = out_ws.cell(row=row_out, column=ci, value=v)
            c.border = thin
            if ci in (4, 7, 10):
                c.alignment = wrap
        row_out += 1

    out_ws.auto_filter.ref = f"A1:K{row_out - 1}"
    out_ws.freeze_panes = "A2"
    out_wb.save(_MISMATCH_XLSX)
    print(f"불일치 {row_out - 2}건 저장: {_MISMATCH_XLSX}")


def main() -> None:
    ap = argparse.ArgumentParser(description="평가 표본 확장: OCR 추출 + 듀얼모델 독립 초안")
    ap.add_argument("--dry", action="store_true", help="API 호출 없이 배선만 점검")
    ap.add_argument("--ocr", action="store_true", help="① 신규 이미지 OCR 문장 추출")
    ap.add_argument("--draft", action="store_true", help="② 모델 2개 독립 초안(①끝난 뒤)")
    ap.add_argument(
        "--export-mismatches", action="store_true",
        help="API 호출 없이 ②산출물에서 불일치 행만 별도 워크북으로 뽑는다",
    )
    ap.add_argument("--limit", type=int, default=None, help="이미지 N장만(스모크 테스트)")
    args = ap.parse_args()

    if args.export_mismatches:
        export_mismatches()
        return

    if args.dry:
        meta = load_expansion_meta()
        print(f"[dry run] 이미지 {len(meta)}장 확인됨(OCR 미실행)")
        for m in meta[:5]:
            print(f"  [{m['nn']}] {m['png']}")
        print("  ...")
        return

    if args.ocr:
        run_ocr(limit=args.limit)
        return

    if args.draft:
        if not _EXPANSION_KEY.exists():
            raise SystemExit("먼저 --ocr로 문장을 추출해야 합니다.")
        meta = json.loads(_EXPANSION_KEY.read_text(encoding="utf-8"))
        if args.limit:
            meta = meta[: args.limit]
        meta = [m for m in meta if m.get("llm_sentences")]
        print(f"초안 대상: 이미지 {len(meta)}장 (문장 있는 것만)")

        vlm_a = get_vlm("openai", model=MODEL_A)
        vlm_b = get_vlm("openai", model=MODEL_B)
        context = build_judgment_context()

        drafts_a: dict[str, dict[int, dict]] = {}
        drafts_b: dict[str, dict[int, dict]] = {}
        for m in meta:
            nn = m["nn"]
            img_path = _IMAGES_DIR / m["png"]
            image_bytes = img_path.read_bytes()
            sentences = m["llm_sentences"]
            print(f"  [{nn}] {m['png']} ({len(sentences)}문장)...", end=" ", flush=True)
            da = draft_one_model_batched(nn, sentences, image_bytes, vlm_a, context)
            db = draft_one_model_batched(nn, sentences, image_bytes, vlm_b, context)
            drafts_a[nn], drafts_b[nn] = da, db
            print(f"{MODEL_A} {len(da)}/{len(sentences)}, {MODEL_B} {len(db)}/{len(sentences)}")

        print(f"\n{MODEL_A} 토큰: {getattr(vlm_a, 'total_tokens', 0)}")
        print(f"{MODEL_B} 토큰: {getattr(vlm_b, 'total_tokens', 0)}")

        write_expansion_xlsx(meta, drafts_a, drafts_b)
        return

    print("--dry, --ocr, --draft 중 하나를 지정하세요.")


if __name__ == "__main__":
    main()
