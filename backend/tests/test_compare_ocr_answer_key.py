"""compare_ocr.load_answer_key()의 헤더 기반 컬럼 조회 유닛테스트.

순수 로직(xlsx 스키마 파싱)이라 실제 API 없이 검증한다. 고정 열번호(G=7·H=8)로
읽다가, 스키마가 다른 정답셋(label_worksheet_reviewed.xlsx)의 240문장이 전부
"제외됨"으로 잘못 처리되는 사고가 실제 있었다(2026-08-18) — 그 회귀 방지용이다.
"""

import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, "scripts")
import compare_ocr  # noqa: E402


def _make_xlsx(path: Path, headers: list[str], rows: list[list]) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "라벨링"
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
    wb.save(path)
    return path


def test_검토필요_제외열_있는_정답셋을_정상_파싱한다(tmp_path):
    p = _make_xlsx(
        tmp_path / "combined.xlsx",
        ["이미지", "상품코드", "문장#", "문장", "판정", "위반유형", "검토필요_사유", "제외사유"],
        [
            ["01", "C1", 1, "문장A", "위반", "1호", "", ""],
            ["01", "C1", 2, "문장B", "검토필요", "", "정보부족형 — 근거", ""],
        ],
    )
    key = compare_ocr.load_answer_key(label_xlsx=p)
    rows = key["01"]
    assert len(rows) == 2
    assert rows[1]["review_kind"] == "정보부족형"


def test_제외열_채워진_행은_안_돌아온다(tmp_path):
    p = _make_xlsx(
        tmp_path / "combined.xlsx",
        ["이미지", "상품코드", "문장#", "문장", "판정", "위반유형", "검토필요_사유", "제외사유"],
        [
            ["01", "C1", 1, "화장품 문장", "위반", "1호", "", ""],
            ["02", "C2", 1, "도구 문장", "대상외", "", "", "대상외상품 — 화장품 아님"],
        ],
    )
    key = compare_ocr.load_answer_key(label_xlsx=p)
    assert "01" in key
    assert "02" not in key


def test_컬럼_이름이_다른_스키마는_고정_열번호로_오해하지_않는다(tmp_path):
    """실제 사고 재현: 8번째 열이 '제외사유'가 아니라 다른 뜻(예: LLM_판정)이면
    전부 제외로 잘못 읽으면 안 된다. 헤더 이름이 없으면 그 기능이 없는 것으로 본다."""
    p = _make_xlsx(
        tmp_path / "reviewed.xlsx",
        ["이미지", "상품코드", "문장#", "문장", "판정", "위반유형", "근거 메모", "LLM_판정", "LLM_위반유형"],
        [
            ["01", "C1", 1, "문장A", "위반", "1호", "메모", "검토필요", "5호"],
            ["01", "C1", 2, "문장B", "대상외", "", "메모", "대상외", ""],
        ],
    )
    key = compare_ocr.load_answer_key(label_xlsx=p)
    assert len(key["01"]) == 2  # 하나도 제외되지 않아야 한다
    assert all(r["review_kind"] == "" for r in key["01"])  # 검토필요_사유 열도 없음


def test_컬럼_없는_정답셋도_기존처럼_동작한다(tmp_path):
    """G·H열 자체가 없는 옛 정답셋(하위호환)."""
    p = _make_xlsx(
        tmp_path / "old.xlsx",
        ["이미지", "상품코드", "문장#", "문장", "판정", "위반유형"],
        [["01", "C1", 1, "문장A", "위반", "1호"]],
    )
    key = compare_ocr.load_answer_key(label_xlsx=p)
    assert len(key["01"]) == 1
    assert key["01"][0]["review_kind"] == ""


def test_L열_비비_최종판단이_판정을_덮어쓴다(tmp_path):
    """기존 동작 회귀 확인: L열(12번째)이 채워지면 그 값이 우선한다."""
    headers = ["이미지", "상품코드", "문장#", "문장", "판정", "위반유형",
               "검토필요_사유", "제외사유", "H2", "H3", "H4", "비비_최종판단"]
    p = _make_xlsx(
        tmp_path / "combined.xlsx", headers,
        [["01", "C1", 1, "문장A", "검토필요", "", "", "", "", "", "", "위반 — 재검토 결과"]],
    )
    key = compare_ocr.load_answer_key(label_xlsx=p)
    assert key["01"][0]["judgment"] == "위반"
