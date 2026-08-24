"""사업자가 올린 엑셀/CSV/TXT에서 전성분+함량을 뽑는다 (create 모드, PM 요청 2026-08-24).

성분을 20~30개 손으로 한 줄씩 치는 게 지옥이라 파일 업로드로 대체한다. 외부 호출이
없는 순수 파싱이라 유닛테스트로 충분히 검증한다.

**행 단위 실패는 조용히 넘기지 않는다.** 몇 행을 왜 건너뛰었는지 `warnings`로 같이
낸다 — 안 그러면 "20개 넣었는데 왜 17개만 들어왔지"를 사용자가 알 방법이 없다.
파일 자체를 못 읽는 실패(헤더를 못 찾음·시트가 없음·손상된 파일)만
`IngredientParseError`로 터뜨린다(API가 4xx로 옮긴다).
"""

import csv
import io
import re

import openpyxl

from barum.models import IngredientAmount

MAX_ROWS = 200

# 헤더를 이 열 수만큼 훑는다. 그 안에 못 찾으면 양식이 아니라고 본다.
_HEADER_SCAN_ROWS = 15

# 참고용 시트(식약처 고시 기준표 등)는 성분 입력이 아니다. 이름에 이 말이 들어가면
# 아예 후보에서 뺀다(팀장 배포 양식에 실제로 있음, PM 확인).
_REFERENCE_SHEET_MARKER = "참고"

# 예시행 판별 마커. "성분명·함량·비고" 어느 칸에 있어도 걸린다.
_EXAMPLE_MARKERS = ("지우고", "채우세요", "예시", "example", "샘플", "입력하세요", "작성하세요")

_AMOUNT_UNIT = r"(?:%|mg/g|IU/g|mg|g|ml|IU)"
# 붙여쓴 한 줄("나이아신아마이드 3%")에서 끝의 함량만 떼는 데 쓴다. 범위 표기
# ("2~5%")까지 받는다.
_TRAILING_AMOUNT_RE = re.compile(
    rf"(\d+(?:[.,]\d+)?(?:\s*~\s*\d+(?:[.,]\d+)?)?\s*{_AMOUNT_UNIT})\s*$",
    re.IGNORECASE,
)


class IngredientParseError(ValueError):
    """파일 자체를 못 읽는 실패. 사용자가 읽을 수 있는 메시지를 담는다."""


def parse_ingredient_upload(ext: str, data: bytes) -> tuple[list[IngredientAmount], list[str]]:
    """확장자에 맞는 파서로 분기한다. (성분 목록, 경고 목록)을 낸다."""
    if ext == ".xlsx":
        return _parse_xlsx(data)
    if ext in (".csv", ".txt"):
        return _parse_delimited_text(data)
    raise IngredientParseError(f"지원하지 않는 확장자입니다: {ext!r}")


# ── xlsx ──────────────────────────────────────────────────────────────────


def _parse_xlsx(data: bytes) -> tuple[list[IngredientAmount], list[str]]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        raise IngredientParseError(f"엑셀 파일을 열 수 없습니다: {type(e).__name__}") from e

    candidates = [wb[name] for name in wb.sheetnames if _REFERENCE_SHEET_MARKER not in name]
    if not candidates:
        raise IngredientParseError(
            "전성분을 담은 시트를 찾지 못했습니다(참고용 시트만 있습니다)."
        )
    # 이름에 "전성분"·"함량"이 다 들어간 시트를 먼저 시도한다. 없으면 나머지를
    # 원래 순서대로 훑어서 실제로 헤더(성분명·함량)가 있는 첫 시트를 쓴다 -
    # "없으면 첫 시트"를 문자 그대로 적용하면 "0_안내" 같은 시트를 잘못 고를 수 있다.
    candidates.sort(key=lambda ws: 0 if ("전성분" in ws.title and "함량" in ws.title) else 1)

    for ws in candidates:
        rows = list(ws.iter_rows(values_only=True))
        header = _find_header(rows)
        if header is not None:
            return _extract_xlsx_rows(rows, header)

    raise IngredientParseError("헤더(성분명·함량)를 찾을 수 없습니다. 양식을 확인해 주세요.")


def _find_header(rows: list[tuple]) -> tuple[int, int, int, int | None] | None:
    """헤더 행을 찾는다. (행 인덱스, 성분명 열, 함량 열, 비고 열) 또는 None.

    열 위치를 이름으로 잡는다 - 순번·비고가 섞여 있어서 "첫 두 열"로 가정하면 깨진다.
    """
    for row_idx, row in enumerate(rows[:_HEADER_SCAN_ROWS]):
        name_col = amount_col = remark_col = None
        for col_idx, cell in enumerate(row):
            text = str(cell).strip() if cell is not None else ""
            if not text:
                continue
            if name_col is None and "성분" in text:
                name_col = col_idx
            elif amount_col is None and "함량" in text:
                amount_col = col_idx
            elif remark_col is None and "비고" in text:
                remark_col = col_idx
        if name_col is not None and amount_col is not None:
            return row_idx, name_col, amount_col, remark_col
    return None


def _cell_text(row: tuple, col: int | None) -> str:
    if col is None or col >= len(row):
        return ""
    value = row[col]
    return "" if value is None else str(value).strip()


def _extract_xlsx_rows(
    rows: list[tuple], header: tuple[int, int, int, int | None]
) -> tuple[list[IngredientAmount], list[str]]:
    header_idx, name_col, amount_col, remark_col = header
    result: list[IngredientAmount] = []
    warnings: list[str] = []
    capped = False

    for offset, row in enumerate(rows[header_idx + 1 :]):
        excel_row = header_idx + 2 + offset  # 1-indexed, 헤더 다음 줄부터
        name = _cell_text(row, name_col)
        amount = _cell_text(row, amount_col)
        remark = _cell_text(row, remark_col)

        if not name and not amount:
            continue  # 빈 행. 양식에 으레 있다 - 경고 없이 넘긴다.
        if _looks_like_example(name, amount, remark):
            continue  # 안내용 예시행. 사용자 입력이 아니라 경고 없이 넘긴다.
        if not amount:
            warnings.append(f"{excel_row}행: 함량이 비어 건너뛰었습니다")
            continue
        if not name:
            warnings.append(f"{excel_row}행: 성분명이 비어 건너뛰었습니다")
            continue
        if len(result) >= MAX_ROWS:
            if not capped:
                warnings.append(f"{MAX_ROWS}행을 넘는 입력은 읽지 않았습니다")
                capped = True
            continue
        result.append(IngredientAmount(name=name, amount=amount))

    return result, warnings


def _looks_like_example(name: str, amount: str, remark: str) -> bool:
    combined = f"{name} {amount} {remark}".lower()
    return any(marker.lower() in combined for marker in _EXAMPLE_MARKERS)


# ── csv / txt ────────────────────────────────────────────────────────────


def _decode(data: bytes) -> str:
    for encoding in ("utf-8-sig", "cp949"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _looks_like_header_line(line: str) -> bool:
    return "성분" in line and "함량" in line


def _split_line(line: str) -> tuple[str, str, bool]:
    """한 줄에서 (성분명, 함량, 성공여부)를 뽑는다.

    **끝의 함량 표기부터 먼저 뗀다.** 콤마·탭 분리를 먼저 하면 "2,500 IU/g"처럼
    함량 안에 천단위 콤마가 있는 경우 그 콤마를 필드 구분자로 오해한다("비타민D 2"
    / "500 IU/g"로 잘못 쪼개졌다, 테스트로 잡음). 끝에서부터 단위가 붙은 숫자를
    먼저 인식하면 콤마·탭이 이름과 함량 사이 구분자든 함량 안의 표기든 상관없이
    맞게 떨어진다.

    함량에 단위가 없어 정규식이 못 잡으면(예: "성분,비고") 콤마·탭 두 칸으로
    폴백한다.
    """
    match = _TRAILING_AMOUNT_RE.search(line)
    if match:
        name = line[: match.start()].rstrip(" \t,").strip()
        if name:
            return name, match.group(1).strip(), True

    for delimiter in (",", "\t"):
        if delimiter in line:
            parts = [p.strip() for p in line.split(delimiter) if p.strip()]
            if len(parts) >= 2:
                return parts[0], parts[1], True

    return "", "", False


def _parse_delimited_text(data: bytes) -> tuple[list[IngredientAmount], list[str]]:
    text = _decode(data)
    result: list[IngredientAmount] = []
    warnings: list[str] = []
    capped = False

    for line_no, raw_line in enumerate(text.splitlines(), start=1):
        line = raw_line.strip()
        if not line:
            continue
        if _looks_like_header_line(line):
            continue

        name, amount, ok = _split_line(line)
        if not ok:
            excerpt = line[:30]
            warnings.append(f"{line_no}행: 형식을 알아볼 수 없어 건너뛰었습니다: {excerpt!r}")
            continue
        if not name or not amount:
            warnings.append(f"{line_no}행: 성분명 또는 함량이 비어 건너뛰었습니다")
            continue
        if len(result) >= MAX_ROWS:
            if not capped:
                warnings.append(f"{MAX_ROWS}행을 넘는 입력은 읽지 않았습니다")
                capped = True
            continue
        result.append(IngredientAmount(name=name, amount=amount))

    return result, warnings
