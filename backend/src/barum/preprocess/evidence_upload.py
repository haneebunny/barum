"""사업자가 올린 엑셀/CSV/TXT에서 실증자료(임상)·설문조사 결과를 뽑는다 (create 모드).

전성분(`ingredient_upload.py`)과 같은 결이다. 외부 호출이 없는 순수 파싱이라
유닛테스트로 검증한다. VLM·LLM은 쓰지 않는다.

전성분과 결정적으로 다른 점이 하나 있다. **헤더가 없는 파일이 온다.** 사업자가
엑셀에서 값만 긁어 붙이면 첫 줄부터 데이터다(제공된 샘플 `설문조사.txt`가
그렇다). 전성분처럼 헤더 이름으로만 열을 잡으면 그런 파일은 통째로 못 읽는다.
그렇다고 열 순서를 고정하면 순서가 다른 파일이 조용히 잘못 들어간다.

그래서 3단으로 잡는다.

1. **헤더 이름** - 첫 행에서 필드 라벨이 둘 이상 잡히면 헤더로 보고 이름으로 매핑
2. **값의 생김새** - `87%`는 수치, `8주`는 기간, `유어랩 피부과학연구소`는 기관
3. **고정 순서 폴백** - 남은 필드만 모델 정의 순서로 채운다

**2·3단으로 읽었으면 반드시 `warnings`를 낸다.** 여기서 열이 한 칸 밀리면
시험기관 자리에 시험기간이 들어간 채로 상세페이지까지 간다. 조용히 넘기면
안 되는 종류의 실패다.

행 단위 실패도 전성분과 같이 `warnings`로 낸다. 파일 자체를 못 읽는 실패만
`EvidenceParseError`로 터뜨린다(API가 4xx로 옮긴다).
"""

import io
import re
from collections.abc import Callable

import openpyxl

from barum.models import ClinicalEvidence, SurveyEvidence

# 실증자료는 제품 하나에 1~3건이 현실이다(성적서 한 건이 2~4개월·수백만원짜리라
# 수십 건이 쌓이지 않는다). 전성분(200행)보다 훨씬 낮게 잡는다.
MAX_ROWS = 50

# 첫 행에서 이만큼 필드가 잡히면 헤더로 본다. 1로 낮추면 데이터 행이 우연히
# 한 단어 걸려 헤더로 오인되고, 그러면 그 행이 통째로 사라진다. 샘플
# `설문조사.txt`는 데이터 한 줄뿐이라 그 오인이 곧 "0건 입력됨"이 된다.
_HEADER_MIN_HITS = 2


class EvidenceParseError(ValueError):
    """파일 자체를 못 읽는 실패. 사용자가 읽을 수 있는 메시지를 담는다."""


# ── 값의 생김새 판정기 ─────────────────────────────────────────────────────
#
# 각 필드가 서로 다른 모양을 갖는다는 점을 이용한다. 헤더가 없어도 값만 보고
# 어느 칸인지 맞출 수 있고, 열 순서가 섞여 있어도 제자리를 찾는다.

# 결과 수치. 끝이 %·배로 끝난다. "4주 후 2.1배"처럼 앞에 기간이 붙어도 잡힌다.
_RATIO_RE = re.compile(r"\d+(?:[.,]\d+)?\s*(?:%|퍼센트|배|점|pt)\s*$", re.IGNORECASE)

# 인원. "150명"·"20명 대상" 둘 다. 임상은 note, 설문은 sample_size로 간다.
_COUNT_RE = re.compile(r"\d+\s*명")

# 시험기간. **단독 표기일 때만** 잡는다(`^...$` 앵커). 앵커가 없으면
# "4주 후 2.1배"가 기간으로 먹혀 value와 자리를 다툰다.
_DURATION_RE = re.compile(r"^\s*(?:약\s*)?\d+(?:[.,]\d+)?\s*(?:주간|주|개월|달|일|년|회|차)\s*$")

# 조사 시기. 연·월 표기.
_DATE_RE = re.compile(r"\d{4}\s*년|\d{1,2}\s*월|\d{4}\s*[-./]\s*\d{1,2}")

_INSTITUTION_MARKERS = (
    "연구소", "연구원", "연구센터", "리서치", "랩", "lab", "센터", "센타",
    "대학", "병원", "의원", "협회", "시험소", "인스티튜트", "institute",
)

_METHOD_MARKERS = (
    "설문", "조사", "인터뷰", "온라인", "오프라인", "면접", "기입",
    "패널", "웹", "모바일", "survey",
)


def _matches_ratio(text: str) -> bool:
    return bool(_RATIO_RE.search(text))


def _matches_count(text: str) -> bool:
    return bool(_COUNT_RE.search(text))


def _matches_duration(text: str) -> bool:
    return bool(_DURATION_RE.match(text))


def _matches_date(text: str) -> bool:
    return bool(_DATE_RE.search(text))


def _has_marker(markers: tuple[str, ...]) -> Callable[[str], bool]:
    def matcher(text: str) -> bool:
        lowered = text.lower()
        return any(marker in lowered for marker in markers)

    return matcher


_matches_institution = _has_marker(_INSTITUTION_MARKERS)
_matches_method = _has_marker(_METHOD_MARKERS)


# ── 필드 명세 ─────────────────────────────────────────────────────────────
#
# `shape_order`의 순서가 곧 생김새 판정의 우선순위다. 앞에 있는 필드가 열을
# 먼저 가져간다. value를 period보다 앞에 두는 이유: "4주 후 2.1배"는 둘 다
# 후보가 될 수 있는데 수치로 읽는 게 맞다(period 판정기가 단독 표기만 잡으니
# 실제로 겹치진 않지만, 순서로도 못 박아둔다).
#
# `header_words`는 헤더 셀에 이 단어가 들어 있으면 그 열로 본다. 샘플
# 헤더가 "결과 수치(value)"처럼 한글 라벨과 영문 필드명을 같이 달고 있어서
# 둘 다 넣는다.

_CLINICAL_FIELDS = ("claim", "value", "institution", "period", "note")
_CLINICAL_LABELS = {
    "claim": "무엇을 개선했는지",
    "value": "결과 수치",
    "institution": "시험기관명",
    "period": "시험기간",
    "note": "피험자 수·조건",
}
_CLINICAL_HEADER_WORDS = {
    "claim": ("claim", "개선", "주장", "무엇을"),
    "value": ("value", "수치", "결과", "값"),
    "institution": ("institution", "기관", "시험소"),
    "period": ("period", "기간"),
    "note": ("note", "피험자", "비고", "부연", "대상자"),
}
_CLINICAL_SHAPE_ORDER = (
    ("value", _matches_ratio),
    ("period", _matches_duration),
    ("note", _matches_count),
    ("institution", _matches_institution),
)
_CLINICAL_REQUIRED = ("claim", "value")

_SURVEY_FIELDS = ("claim", "value", "sample_size", "institution", "period", "method")
_SURVEY_LABELS = {
    "claim": "무엇에 대한 응답인지",
    "value": "결과 수치",
    "sample_size": "표본 수",
    "institution": "조사기관명",
    "period": "조사 시기",
    "method": "조사 방법",
}
_SURVEY_HEADER_WORDS = {
    "claim": ("claim", "응답", "무엇에", "주장"),
    "value": ("value", "수치", "결과", "값"),
    "sample_size": ("sample_size", "표본", "응답자", "인원", "샘플"),
    "institution": ("institution", "기관"),
    "period": ("period", "시기", "기간", "일자"),
    "method": ("method", "방법", "방식"),
}
# institution을 method보다 앞에 둔다. "한국갤럽조사연구소"처럼 기관명에 "조사"가
# 들어가면 둘 다 걸리는데, 열은 먼저 잡은 필드가 가져간다.
_SURVEY_SHAPE_ORDER = (
    ("value", _matches_ratio),
    ("sample_size", _matches_count),
    ("period", _matches_date),
    ("institution", _matches_institution),
    ("method", _matches_method),
)
# 설문은 6칸이 전부 필수다(`SurveyEvidence` docstring 참조). 다만 파서에서
# 행을 버리진 않는다 - 프론트가 `isSurveyEvidenceComplete`로 걸러내고, 여기선
# 빈 칸이 있다는 사실만 경고로 알린다.
_SURVEY_REQUIRED = ("claim", "value")


# ── 공개 API ──────────────────────────────────────────────────────────────


def parse_clinical_upload(ext: str, data: bytes) -> tuple[list[ClinicalEvidence], list[str]]:
    """실증자료(임상) 파일을 파싱한다. (자료 목록, 경고 목록)을 낸다."""
    rows, warnings = _parse(
        ext,
        data,
        fields=_CLINICAL_FIELDS,
        labels=_CLINICAL_LABELS,
        header_words=_CLINICAL_HEADER_WORDS,
        shape_order=_CLINICAL_SHAPE_ORDER,
        required=_CLINICAL_REQUIRED,
    )
    return [ClinicalEvidence(**row) for row in rows], warnings


def parse_survey_upload(ext: str, data: bytes) -> tuple[list[SurveyEvidence], list[str]]:
    """설문조사 결과 파일을 파싱한다. (설문 목록, 경고 목록)을 낸다.

    6칸이 다 안 차도 행을 버리지 않는다. 사용자가 폼에서 마저 채울 수 있어야
    하고, 어느 행이 왜 미완인지는 경고로 알린다.
    """
    rows, warnings = _parse(
        ext,
        data,
        fields=_SURVEY_FIELDS,
        labels=_SURVEY_LABELS,
        header_words=_SURVEY_HEADER_WORDS,
        shape_order=_SURVEY_SHAPE_ORDER,
        required=_SURVEY_REQUIRED,
    )
    result = [SurveyEvidence(**row) for row in rows]
    for idx, row in enumerate(result, start=1):
        missing = [_SURVEY_LABELS[f] for f in _SURVEY_FIELDS if not getattr(row, f).strip()]
        if missing:
            warnings.append(
                f"{idx}번째 설문: {'·'.join(missing)}이(가) 비어 있습니다. "
                "6개 항목을 모두 채워야 생성에 쓰입니다."
            )
    return result, warnings


# ── 파일 → 행렬 ───────────────────────────────────────────────────────────


def _cells_from_file(ext: str, data: bytes) -> list[list[str]]:
    """확장자에 맞게 읽어 문자열 행렬로 만든다. 빈 행은 여기서 걸러낸다."""
    if ext == ".xlsx":
        return _cells_from_xlsx(data)
    if ext in (".csv", ".txt"):
        return _cells_from_text(data)
    raise EvidenceParseError(f"지원하지 않는 확장자입니다: {ext!r}")


def _cells_from_xlsx(data: bytes) -> list[list[str]]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        raise EvidenceParseError(f"엑셀 파일을 열 수 없습니다: {type(e).__name__}") from e

    for ws in (wb[name] for name in wb.sheetnames):
        rows = [
            [("" if cell is None else str(cell).strip()) for cell in row]
            for row in ws.iter_rows(values_only=True)
        ]
        rows = [row for row in rows if any(row)]
        if rows:
            return rows
    raise EvidenceParseError("내용이 있는 시트를 찾지 못했습니다.")


def _decode(data: bytes) -> str:
    for encoding in ("utf-8-sig", "cp949"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _cells_from_text(data: bytes) -> list[list[str]]:
    """탭 우선, 없으면 콤마로 쪼갠다.

    탭을 먼저 보는 이유: 엑셀·구글시트에서 복사해 붙이면 탭으로 떨어지는데,
    그 안의 값에 "2,500명"처럼 천단위 콤마가 섞여 있으면 콤마부터 보다가
    한 칸이 두 칸으로 쪼개진다.
    """
    rows: list[list[str]] = []
    for raw_line in _decode(data).splitlines():
        # 줄 끝만 떼고 앞은 그대로 둔다. `line.strip()`으로 시작하면 첫 칸이
        # 빈 행("\t2.1배\t...")에서 **선행 탭이 지워져 열이 통째로 한 칸
        # 밀린다.** 그러면 빈 필수항목이 안 보이고 엉뚱한 값이 실린다.
        line = raw_line.rstrip("\r\n")
        if not line.strip():
            continue
        delimiter = "\t" if "\t" in line else ","
        cells = [c.strip() for c in line.split(delimiter)]
        if any(cells):
            rows.append(cells)
    return rows


# ── 열 → 필드 매핑 ────────────────────────────────────────────────────────


def _match_header(
    header: list[str], fields: tuple[str, ...], header_words: dict[str, tuple[str, ...]]
) -> dict[str, int]:
    """헤더 행에서 (필드 -> 열 인덱스)를 찾는다. 한 열은 한 필드만 가져간다."""
    assigned: dict[str, int] = {}
    used: set[int] = set()
    for field in fields:
        words = header_words[field]
        for col_idx, cell in enumerate(header):
            if col_idx in used:
                continue
            lowered = cell.lower()
            if lowered and any(word in lowered for word in words):
                assigned[field] = col_idx
                used.add(col_idx)
                break
    return assigned


def _column_matches(values: list[str], matcher: Callable[[str], bool]) -> bool:
    """그 열의 값 과반이 이 모양이면 참.

    행 단위가 아니라 열 단위로 본다. 임상처럼 행이 여러 개면 한 행이 애매해도
    나머지가 받쳐준다.
    """
    filled = [v for v in values if v]
    if not filled:
        return False
    hits = sum(1 for v in filled if matcher(v))
    return hits * 2 > len(filled)


def _match_shape(
    columns: list[list[str]],
    shape_order: tuple[tuple[str, Callable[[str], bool]], ...],
    already_used: set[int],
) -> dict[str, int]:
    """값의 생김새로 (필드 -> 열 인덱스)를 찾는다."""
    assigned: dict[str, int] = {}
    used = set(already_used)
    for field, matcher in shape_order:
        for col_idx, values in enumerate(columns):
            if col_idx in used:
                continue
            if _column_matches(values, matcher):
                assigned[field] = col_idx
                used.add(col_idx)
                break
    return assigned


def _resolve_columns(
    rows: list[list[str]],
    fields: tuple[str, ...],
    labels: dict[str, str],
    header_words: dict[str, tuple[str, ...]],
    shape_order: tuple[tuple[str, Callable[[str], bool]], ...],
) -> tuple[dict[str, int], list[list[str]], list[str]]:
    """3단으로 열을 잡는다. (필드->열, 데이터 행들, 경고)를 낸다."""
    warnings: list[str] = []

    # 1단: 헤더
    header_map = _match_header(rows[0], fields, header_words)
    header_used = len(header_map) >= _HEADER_MIN_HITS
    if header_used:
        data_rows = rows[1:]
        assigned = header_map
        if not data_rows:
            raise EvidenceParseError("헤더만 있고 내용이 없습니다.")
    else:
        data_rows = rows
        assigned = {}

    if not data_rows:
        raise EvidenceParseError("읽을 내용이 없습니다.")

    width = max(len(row) for row in data_rows)
    columns = [[_cell(row, i) for row in data_rows] for i in range(width)]

    # 2단: 값의 생김새. 헤더가 못 채운 필드만 채운다.
    remaining_shape = tuple((f, m) for f, m in shape_order if f not in assigned)
    shape_map = _match_shape(columns, remaining_shape, set(assigned.values()))
    assigned.update(shape_map)

    # 3단: 고정 순서 폴백. 남은 필드를 남은 열에 정의 순서대로 꽂는다.
    leftover_fields = [f for f in fields if f not in assigned]
    leftover_cols = [i for i in range(width) if i not in set(assigned.values())]
    fallback: list[str] = []
    for field, col_idx in zip(leftover_fields, leftover_cols):
        assigned[field] = col_idx
        fallback.append(f"{col_idx + 1}열={labels[field]}")

    # 헤더로 확정한 열은 안전하다. **그 외의 방법으로 잡은 열은 전부 알린다.**
    # 형태 판정(2단)도 추측이고, 고정 순서(3단)는 더 그렇다. 여기서 한 칸 밀리면
    # 시험기관 자리에 시험기간이 들어간 채로 상세페이지까지 간다.
    if not header_used:
        readout = "·".join(
            f"{assigned[f] + 1}열={labels[f]}" for f in fields if f in assigned
        )
        # 폴백으로 잡힌 열을 따로 나열하지 않는다. `claim`은 고유한 형태가 없어
        # 설계상 늘 "나머지"로 떨어지는데, 그때마다 "판단하지 못했다"고 알리면
        # 매번 뜨는 경고가 되어 정작 봐야 할 때 안 읽힌다. 배치 전체를 보여주는
        # 것으로 확인에는 충분하다.
        warnings.append(
            f"헤더가 없어 값의 형태로 열을 읽었습니다: {readout}. "
            "값이 제자리에 들어갔는지 확인해 주세요."
        )
    else:
        guessed = [f"{assigned[f] + 1}열={labels[f]}" for f in fields if f in shape_map]
        parts = []
        if guessed:
            parts.append(f"값의 형태로 읽은 열 {'·'.join(guessed)}")
        if fallback:
            parts.append(f"기본 순서로 읽은 열 {'·'.join(fallback)}")
        if parts:
            warnings.append(
                f"헤더에 이름이 없는 항목이 있습니다: {', '.join(parts)}. "
                "값이 제자리에 들어갔는지 확인해 주세요."
            )

    unmapped = [labels[f] for f in fields if f not in assigned]
    if unmapped:
        warnings.append(f"열이 부족해 채우지 못한 항목: {'·'.join(unmapped)}")

    return assigned, data_rows, warnings


def _cell(row: list[str], col_idx: int) -> str:
    return row[col_idx] if col_idx < len(row) else ""


# ── 본체 ──────────────────────────────────────────────────────────────────


def _parse(
    ext: str,
    data: bytes,
    *,
    fields: tuple[str, ...],
    labels: dict[str, str],
    header_words: dict[str, tuple[str, ...]],
    shape_order: tuple[tuple[str, Callable[[str], bool]], ...],
    required: tuple[str, ...],
) -> tuple[list[dict[str, str]], list[str]]:
    rows = _cells_from_file(ext, data)
    if not rows:
        raise EvidenceParseError("내용이 없는 파일입니다.")

    assigned, data_rows, warnings = _resolve_columns(
        rows, fields, labels, header_words, shape_order
    )

    result: list[dict[str, str]] = []
    capped = False
    for offset, row in enumerate(data_rows):
        line_no = offset + 1 + (len(rows) - len(data_rows))  # 1-indexed, 헤더 포함
        values = {f: (_cell(row, assigned[f]) if f in assigned else "") for f in fields}

        missing = [labels[f] for f in required if not values[f]]
        if missing:
            warnings.append(f"{line_no}행: {'·'.join(missing)}이(가) 비어 건너뛰었습니다")
            continue
        if len(result) >= MAX_ROWS:
            if not capped:
                warnings.append(f"{MAX_ROWS}행을 넘는 입력은 읽지 않았습니다")
                capped = True
            continue
        result.append(values)

    return result, warnings
